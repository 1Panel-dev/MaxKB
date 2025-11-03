def parse_excel_to_csv_optimized(
        file_url="https://oss.goservice.cn/sparkone/2025/08/20/f06339d1d12a4a2d97f694ec18a5ba23.xlsx", use_cache=True):
    """
    从 OSS 存储下载并解析 Excel 文件，输出 CSV 格式的简化结果
    优化的缓存策略：先缓存原始文件，再缓存解析结果

    参数:
    - file_url: Excel 文件的 URL 地址
    - use_cache: 是否使用缓存（默认 True）

    返回:
    - 简化的 CSV 格式数据，便于大模型理解
    """
    import os
    import hashlib
    import pickle
    import requests
    from io import BytesIO, StringIO
    from urllib.parse import urlparse
    import time

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "error": "缺少 openpyxl 库，请安装: pip install openpyxl",
            "status": "依赖缺失"
        }

    try:
        # 缓存目录配置
        cache_dir = "/opt/maxkb/cache"
        os.makedirs(cache_dir, exist_ok=True)

        # 从URL提取文件名
        parsed_url = urlparse(file_url)
        original_filename = os.path.basename(parsed_url.path)
        if not original_filename or not original_filename.endswith(('.xlsx', '.xls')):
            # 如果无法从URL提取文件名，使用hash值
            url_hash = hashlib.md5(file_url.encode()).hexdigest()[:8]
            original_filename = f"excel_file_{url_hash}.xlsx"

        # 缓存文件路径
        cached_excel_file = os.path.join(cache_dir, original_filename)
        cached_result_file = os.path.join(cache_dir, f"{original_filename}_parsed.pkl")

        print(f"📁 缓存文件: {original_filename}")

        # 检查是否存在缓存的解析结果
        if use_cache and os.path.exists(cached_result_file):
            try:
                print("📋 从缓存加载解析结果...")
                with open(cached_result_file, 'rb') as f:
                    cached_result = pickle.load(f)
                    # 添加缓存信息
                    cached_result["cache_info"] = {
                        "from_cache": True,
                        "cached_file": original_filename,
                        "cache_time": time.ctime(os.path.getmtime(cached_result_file))
                    }
                    return cached_result
            except Exception as e:
                print(f"⚠️ 缓存文件损坏，将重新解析: {e}")

        # 检查是否存在缓存的Excel文件
        excel_data = None
        if use_cache and os.path.exists(cached_excel_file):
            try:
                print(f"📂 从缓存读取Excel文件: {original_filename}")
                with open(cached_excel_file, 'rb') as f:
                    excel_data = f.read()
                print(f"✅ 缓存文件读取成功，大小: {len(excel_data)} 字节")
            except Exception as e:
                print(f"⚠️ 缓存Excel文件读取失败，将重新下载: {e}")
                excel_data = None

        # 如果没有缓存文件或读取失败，则下载
        if excel_data is None:
            print(f"🌐 开始下载文件: {file_url}")
            response = requests.get(file_url, timeout=30)
            response.raise_for_status()
            excel_data = response.content
            print(f"✅ 文件下载成功，大小: {len(excel_data)} 字节")

            # 保存Excel文件到缓存
            if use_cache:
                try:
                    with open(cached_excel_file, 'wb') as f:
                        f.write(excel_data)
                    print(f"💾 Excel文件已缓存: {cached_excel_file}")
                except Exception as e:
                    print(f"⚠️ 保存Excel缓存失败: {e}")

        # 解析Excel文件
        excel_buffer = BytesIO(excel_data)
        workbook = load_workbook(excel_buffer, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames

        print(f"📊 发现 {len(sheet_names)} 个工作表: {sheet_names}")

        # 简化的解析结果
        result = {
            "file_info": {
                "filename": original_filename,
                "file_size": len(excel_data),
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
                "source_url": file_url
            },
            "sheets": {}
        }

        # 解析每个 sheet 为 CSV 格式
        for sheet_name in sheet_names:
            print(f"🔍 正在解析工作表: {sheet_name}")

            try:
                worksheet = workbook[sheet_name]
                max_row = worksheet.max_row
                max_col = worksheet.max_column

                if max_row <= 1:  # 只有标题行或空表
                    result["sheets"][sheet_name] = {
                        "info": f"工作表 '{sheet_name}' 为空或只有标题行",
                        "csv_data": ""
                    }
                    continue

                # 生成 CSV 格式的数据
                csv_output = StringIO()

                # 限制最大行数（包括标题行）
                max_rows_to_process = min(max_row, 101)  # 最多100行数据 + 1行标题

                # 写入数据行
                for row_num in range(1, max_rows_to_process + 1):
                    row_data = []
                    for col_num in range(1, max_col + 1):
                        cell_value = worksheet.cell(row=row_num, column=col_num).value

                        # 处理单元格值
                        if cell_value is None:
                            row_data.append("")
                        elif isinstance(cell_value, str):
                            # 处理 CSV 中的特殊字符
                            cleaned_value = cell_value.replace('"', '""').replace('\n', ' ').replace('\r', ' ').strip()
                            if ',' in cleaned_value or '"' in cleaned_value or '\n' in str(cell_value):
                                row_data.append(f'"{cleaned_value}"')
                            else:
                                row_data.append(cleaned_value)
                        else:
                            row_data.append(str(cell_value))

                    csv_output.write(','.join(row_data) + '\n')

                csv_data = csv_output.getvalue()
                csv_output.close()

                # 统计信息
                data_rows = max_row - 1
                info_text = f"工作表 '{sheet_name}': {data_rows} 行数据, {max_col} 列"
                if data_rows > 100:
                    info_text += " (显示前100行)"

                result["sheets"][sheet_name] = {
                    "info": info_text,
                    "rows": data_rows,
                    "columns": max_col,
                    "csv_data": csv_data.strip()
                }

                print(f"✅ 工作表 '{sheet_name}' 转换完成: {data_rows} 行 x {max_col} 列")

            except Exception as e:
                print(f"❌ 解析工作表 '{sheet_name}' 时出错: {str(e)}")
                result["sheets"][sheet_name] = {
                    "info": f"工作表 '{sheet_name}' 解析失败: {str(e)}",
                    "csv_data": "",
                    "error": str(e)
                }

        # 生成总结
        successful_sheets = len([s for s in result["sheets"].values() if s.get("csv_data")])
        total_rows = sum(s.get("rows", 0) for s in result["sheets"].values() if isinstance(s.get("rows"), int))

        result["summary"] = {
            "status": "成功",
            "processed_sheets": successful_sheets,
            "total_sheets": len(sheet_names),
            "total_data_rows": total_rows,
            "description": f"成功解析 {successful_sheets}/{len(sheet_names)} 个工作表，共 {total_rows} 行数据"
        }

        result["cache_info"] = {
            "from_cache": False,
            "cached_file": original_filename,
            "cache_time": time.ctime()
        }

        # 保存解析结果到缓存
        if use_cache:
            try:
                with open(cached_result_file, 'wb') as f:
                    pickle.dump(result, f)
                print(f"💾 解析结果已缓存: {cached_result_file}")
            except Exception as e:
                print(f"⚠️ 保存解析结果缓存失败: {e}")

        print("🎉 Excel 文件解析完成!")
        return result

    except requests.RequestException as e:
        return {
            "error": f"下载文件失败: {str(e)}",
            "status": "下载失败",
            "file_url": file_url
        }

    except Exception as e:
        return {
            "error": f"解析文件失败: {str(e)}",
            "status": "解析失败",
            "file_url": file_url
        }