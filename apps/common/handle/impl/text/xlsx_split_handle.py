# coding=utf-8
"""
    @project: maxkb
    @Author：虎
    @file： xlsx_parse_qa_handle.py
    @date：2024/5/21 14:59
    @desc:
"""
import io
import traceback
from typing import List

import openpyxl
from openpyxl import load_workbook

from common.handle.base_split_handle import BaseSplitHandle
from common.handle.impl.common_handle import xlsx_embed_cells_images
from common.handle.impl.text.excel_kv_common import (
    make_kv_paragraph,
    normalize_headers,
)
from common.utils.logger import maxkb_logger

splitter = '\n`-----------------------------------`\n'


def _row_values_with_merge(sheet, row_idx: int, merged_ranges) -> list:
    """读取指定行的单元格值，对合并单元格做左上角值传播。

    openpyxl 中合并单元格只有左上角单元有值，其余为 None；
    这里把这些 None 替换成合并范围左上角的值，便于按行序列化。
    """
    row_cells = list(sheet[row_idx])
    values = []
    for cell in row_cells:
        val = cell.value
        if val is None:
            for rng in merged_ranges:
                if cell.coordinate in rng:
                    val = sheet[rng.min_row][rng.min_col - 1].value
                    break
        values.append(val)
    return values


def handle_sheet(file_name, sheet, image_dict, limit: int):
    """按行 KV 序列化整张 sheet。每一非空行 → 一个 paragraph。

    file_name 实际上是 sheet 内容描述用的"上下文名称"，
    在外层调用时已根据 sheet 数量决定是用文件名还是 sheet 名作为 context。
    """
    paragraphs = []
    result = {'name': file_name, 'content': paragraphs}
    try:
        if sheet.max_row is None or sheet.max_row < 1 or sheet.max_column is None or sheet.max_column < 1:
            return result
        merged_ranges = list(sheet.merged_cells.ranges)
        # 第一行作为表头
        header_values = _row_values_with_merge(sheet, 1, merged_ranges)
        headers = normalize_headers(header_values)
        if not headers:
            return result
        # 从第 2 行开始作为数据
        for row_idx in range(2, sheet.max_row + 1):
            row_values = _row_values_with_merge(sheet, row_idx, merged_ranges)
            paragraph = make_kv_paragraph(
                file_name=file_name,
                sheet_name=sheet.title,
                row_idx=row_idx,
                headers=headers,
                row_values=row_values,
                image_dict=image_dict,
                limit=limit,
            )
            if paragraph is not None:
                paragraphs.append(paragraph)
    except Exception as e:
        maxkb_logger.error(f"Error parsing XLSX sheet {sheet.title}: {e}, {traceback.format_exc()}")
    return result


class XlsxSplitHandle(BaseSplitHandle):
    def fill_merged_cells(self, sheet, image_dict):
        data = []

        # 获取第一行作为标题行
        headers = []
        for idx, cell in enumerate(sheet[1]):
            if cell.value is None:
                headers.append(' ' * (idx + 1))
            else:
                headers.append(cell.value)

        # 从第二行开始遍历每一行
        for row in sheet.iter_rows(min_row=2, values_only=False):
            row_data = {}
            for col_idx, cell in enumerate(row):
                cell_value = cell.value

                # 如果单元格为空，并且该单元格在合并单元格内，获取合并单元格的值
                if cell_value is None:
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            cell_value = sheet[merged_range.min_row][merged_range.min_col - 1].value
                            break

                image = image_dict.get(cell_value, None)
                if image is not None:
                    cell_value = f'![](./oss/file/{image.id})'

                # 使用标题作为键，单元格的值作为值存入字典
                row_data[headers[col_idx]] = cell_value
            data.append(row_data)

        return data

    def handle(self, file, pattern_list: List, with_filter: bool, limit: int, get_buffer, save_image):
        buffer = get_buffer(file)
        try:
            if type(limit) is str:
                limit = int(limit)
            workbook = openpyxl.load_workbook(io.BytesIO(buffer))
            try:
                image_dict: dict = xlsx_embed_cells_images(io.BytesIO(buffer))
                save_image([item for item in image_dict.values()])
            except Exception as e:
                image_dict = {}
            worksheets = workbook.worksheets
            worksheets_size = len(worksheets)
            results = []
            for sheet in worksheets:
                # paragraph 内的「文件：xxx」始终用真实文件名，与文档拆分名解耦
                sheet_result = handle_sheet(file.name, sheet, image_dict, limit)
                # 多 sheet 工作簿时，仍按 sheet 名拆分为多文档（保留原行为）
                if worksheets_size == 1 and sheet.title == 'Sheet1':
                    sheet_result['name'] = file.name
                else:
                    sheet_result['name'] = sheet.title
                results.append(sheet_result)
            return [r for r in results if r is not None]
        except Exception as e:
            maxkb_logger.error(f"Error processing XLSX file {file.name}: {e}, {traceback.format_exc()}")
            return [{'name': file.name, 'content': []}]

    def get_content(self, file, save_image):
        try:
            # 加载 Excel 文件
            workbook = load_workbook(file)
            try:
                image_dict: dict = xlsx_embed_cells_images(file)
                if len(image_dict) > 0:
                    save_image(image_dict.values())
            except Exception as e:
                maxkb_logger.error(f'Exception: {e}')
                image_dict = {}
            md_tables = ''
            # 遍历所有工作表
            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname]
                rows = self.fill_merged_cells(sheet, image_dict)
                if len(rows) == 0:
                    continue

                # 添加 sheet 名称作为标题
                md_tables += f'## {sheetname}\n\n'

                # 提取表头和内容
                headers = [f"{key}" for key, value in rows[0].items()]

                # 构建 Markdown 表格
                md_table = '| ' + ' | '.join(headers) + ' |\n'
                md_table += '| ' + ' | '.join(['---'] * len(headers)) + ' |\n'
                for row in rows:
                    r = [self._escape_cell_content(value) for key, value in row.items()]
                    md_table += '| ' + ' | '.join(r) + ' |\n'

                md_tables += md_table + '\n\n'

            return md_tables
        except Exception as e:
            maxkb_logger.error(f'excel split handle error: {e}')
            return f'error: {e}'

    def _escape_cell_content(self, cell_value):
        """转义单元格内容,避免破坏 Markdown 表格结构"""
        if cell_value is None:
            return ''

        cell_str = str(cell_value)

        # 替换换行符为 <br>
        cell_str = cell_str.replace('\n', '<br>')

        # 转义管道符 | 为 HTML 实体
        cell_str = cell_str.replace('|', '&#124;')

        # 如果内容包含反引号,需要转义
        if '`' in cell_str:
            cell_str = cell_str.replace('`', '&#96;')

        return cell_str

    def support(self, file, get_buffer):
        file_name: str = file.name.lower()
        if file_name.endswith(".xlsx"):
            return True
        return False
