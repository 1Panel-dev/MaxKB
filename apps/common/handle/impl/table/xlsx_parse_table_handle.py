# coding=utf-8
import io
import logging

from openpyxl import load_workbook

from common.handle.base_parse_table_handle import BaseParseTableHandle
from common.handle.impl.tools import xlsx_embed_cells_images

max_kb = logging.getLogger("max_kb")


class XlsxSplitHandle(BaseParseTableHandle):
    def support(self, file, get_buffer):
        file_name: str = file.name.lower()
        if file_name.endswith('.xlsx'):
            return True
        return False

    def fill_merged_cells(self, sheet, image_dict):
        data = []
        # 从第二行开始遍历每一行
        for row in sheet.iter_rows(values_only=False):
            row_data = []
            for col_idx, cell in enumerate(row):
                cell_value = cell.value
                image = image_dict.get(cell_value, None)
                if image is not None:
                    cell_value = f'![](/api/image/{image.id})'

                # 使用标题作为键，单元格的值作为值存入字典
                row_data.insert(col_idx, cell_value)
            data.append(row_data)

        for merged_range in sheet.merged_cells.ranges:
            cell_value = data[merged_range.min_row - 1][merged_range.min_col - 1]
            for row_index in range(merged_range.min_row, merged_range.max_row + 1):
                for col_index in range(merged_range.min_col, merged_range.max_col + 1):
                    data[row_index - 1][col_index - 1] = cell_value
        return data

    def handle(self, file, get_buffer, save_image):
        buffer = get_buffer(file)
        try:
            wb = load_workbook(io.BytesIO(buffer))
            try:
                image_dict: dict = xlsx_embed_cells_images(io.BytesIO(buffer))
                save_image([item for item in image_dict.values()])
            except Exception as e:
                image_dict = {}
            result = []
            for sheetname in wb.sheetnames:
                paragraphs = []
                ws = wb[sheetname]
                data = self.fill_merged_cells(ws, image_dict)
                if len(data) >= 2:
                    head_list = data[0]
                    for row_index in range(1, len(data)):
                        row_output = "; ".join(
                            [f"{head_list[col_index]}: {data[row_index][col_index]}" for col_index in
                             range(0, len(data[row_index]))])
                        paragraphs.append({'title': '', 'content': row_output})

                result.append({'name': sheetname, 'paragraphs': paragraphs})

        except BaseException as e:
            max_kb.error(f'excel split handle error: {e}')
            return [{'name': file.name, 'paragraphs': []}]
        return result

    def get_content(self, file, save_image):
        try:
            # 加载 Excel 文件
            workbook = load_workbook(file)
            try:
                image_dict: dict = xlsx_embed_cells_images(file)
                if len(image_dict) > 0:
                    save_image(image_dict.values())
            except Exception as e:
                print(f'{e}')
                image_dict = {}
            md_tables = ''
            # 如果未指定 sheet_name，则使用第一个工作表
            for sheetname in workbook.sheetnames:
                sheet = workbook[sheetname] if sheetname else workbook.active
                data = self.fill_merged_cells(sheet, image_dict)
                if len(data) == 0:
                    continue
                # 提取表头和内容

                headers = [f"{value}" for value in data[0]]

                # 构建 Markdown 表格
                md_table = '| ' + ' | '.join(headers) + ' |\n'
                md_table += '| ' + ' | '.join(['---'] * len(headers)) + ' |\n'
                for row_index in range(1, len(data)):
                    r = [f'{value}' for value in data[row_index]]
                    md_table += '| ' + ' | '.join(
                        [str(cell).replace('\n', '<br>') if cell is not None else '' for cell in r]) + ' |\n'

                md_tables += md_table + '\n\n'

                md_tables = md_tables.replace('/api/image/', '/api/file/')
            return md_tables
        except Exception as e:
            max_kb.error(f'excel split handle error: {e}')
            return f'error: {e}'
