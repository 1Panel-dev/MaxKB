# coding=utf-8
import io
import logging

from openpyxl import load_workbook

from common.handle.base_parse_table_handle import BaseParseTableHandle
from common.handle.impl.tools import xlsx_embed_cells_images

max_kb = logging.getLogger("max_kb")


class ExcelSplitHandle(BaseParseTableHandle):
    def support(self, file, get_buffer):
        file_name: str = file.name.lower()
        if file_name.endswith('.xls') or file_name.endswith('.xlsx'):
            return True
        return False

    def handle(self, file, get_buffer, save_image):
        buffer = get_buffer(file)
        try:
            wb = load_workbook(io.BytesIO(buffer))
            try:
                image_dict: dict = xlsx_embed_cells_images(io.BytesIO(buffer))
                save_image([item for item in image_dict.values()])
            except Exception as e:
                image_dict = {}
            result = []
            for sheetname in wb.sheetnames:
                paragraphs = []
                ws = wb[sheetname]
                rows = list(ws.rows)
                if not rows: continue
                ti = list(rows[0])
                for r in list(rows[1:]):
                    title = []
                    l = []
                    for i, c in enumerate(r):
                        if not c.value:
                            continue
                        t = str(ti[i].value) if i < len(ti) else ""
                        title.append(t)
                        content = str(c.value)
                        image = image_dict.get(content, None)
                        if image is not None:
                            content = f'![](/api/image/{image.id})'
                        t += (": " if t else "") + content
                        l.append(t)
                    l = "; ".join(l)
                    if sheetname.lower().find("sheet") < 0:
                        l += " ——" + sheetname
                    paragraphs.append({'title': '', 'content': l})
                result.append({'name': sheetname, 'paragraphs': paragraphs})

        except BaseException as e:
            max_kb.error(f'excel split handle error: {e}')
            return [{'name': file.name, 'paragraphs': []}]
        return result
