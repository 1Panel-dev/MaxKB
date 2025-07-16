# coding=utf-8
"""
    @project: MaxKB
    @Author：虎虎
    @file： application_chat.py
    @date：2025/6/10 11:06
    @desc:
"""
import datetime
import os
import re
from io import BytesIO
from typing import Dict

import openpyxl
import pytz
from django.core import validators
from django.db import models
from django.db.models import QuerySet, Q
from django.http import StreamingHttpResponse
from django.utils.translation import gettext_lazy as _, gettext
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from rest_framework import serializers

from application.models import Chat, Application, ChatRecord
from common.db.search import get_dynamics_model, native_search, native_page_search
from common.exception.app_exception import AppApiException
from common.utils.common import get_file_content
from maxkb.conf import PROJECT_DIR
from maxkb.settings import TIME_ZONE, edition


class ApplicationChatResponseSerializers(serializers.Serializer):
    id = serializers.UUIDField(required=True, label=_("chat id"))
    abstract = serializers.CharField(required=True, label=_("summary"))
    chat_user_id = serializers.UUIDField(required=True, label=_("Chat User ID"))
    chat_user_type = serializers.CharField(required=True, label=_("Chat User Type"))
    is_deleted = serializers.BooleanField(required=True, label=_("Is delete"))
    application_id = serializers.UUIDField(required=True, label=_("Application ID"))
    chat_record_count = serializers.IntegerField(required=True, label=_("Number of conversations"))
    star_num = serializers.IntegerField(required=True, label=_("Number of Likes"))
    trample_num = serializers.IntegerField(required=True, label=_("Number of thumbs-downs"))
    mark_sum = serializers.IntegerField(required=True, label=_("Number of tags"))


class ApplicationChatRecordExportRequest(serializers.Serializer):
    select_ids = serializers.ListField(required=True, label=_("Chat ID List"),
                                       child=serializers.UUIDField(required=True, label=_("Chat ID")))


class ApplicationChatQuerySerializers(serializers.Serializer):
    workspace_id = serializers.CharField(required=False, allow_null=True, allow_blank=True, label=_("Workspace ID"))
    abstract = serializers.CharField(required=False, allow_blank=True, allow_null=True, label=_("summary"))
    start_time = serializers.DateField(format='%Y-%m-%d', label=_("Start time"))
    end_time = serializers.DateField(format='%Y-%m-%d', label=_("End time"))
    application_id = serializers.UUIDField(required=True, label=_("Application ID"))
    min_star = serializers.IntegerField(required=False, min_value=0,
                                        label=_("Minimum number of likes"))
    min_trample = serializers.IntegerField(required=False, min_value=0,
                                           label=_("Minimum number of clicks"))
    comparer = serializers.CharField(required=False, label=_("Comparator"), validators=[
        validators.RegexValidator(regex=re.compile("^and|or$"),
                                  message=_("Only supports and|or"), code=500)
    ])

    def is_valid(self, *, raise_exception=False):
        super().is_valid(raise_exception=True)
        workspace_id = self.data.get('workspace_id')
        query_set = QuerySet(Application).filter(id=self.data.get('application_id'))
        if workspace_id:
            query_set = query_set.filter(workspace_id=workspace_id)
        if not query_set.exists():
            raise AppApiException(500, _('Application id does not exist'))

    def get_end_time(self):
        return datetime.datetime.combine(
            datetime.datetime.strptime(self.data.get('end_time'), '%Y-%m-%d'),
            datetime.datetime.max.time())

    def get_start_time(self):
        return self.data.get('start_time')

    def get_query_set(self, select_ids=None):
        end_time = self.get_end_time()
        start_time = self.get_start_time()
        query_set = QuerySet(model=get_dynamics_model(
            {'application_chat.application_id': models.CharField(),
             'application_chat.abstract': models.CharField(),
             "star_num": models.IntegerField(),
             'trample_num': models.IntegerField(),
             'comparer': models.CharField(),
             'application_chat.update_time': models.DateTimeField(),
             'application_chat.id': models.UUIDField(), }))

        base_query_dict = {'application_chat.application_id': self.data.get("application_id"),
                           'application_chat.update_time__gte': start_time,
                           'application_chat.update_time__lte': end_time,
                           }
        if 'abstract' in self.data and self.data.get('abstract') is not None:
            base_query_dict['application_chat.abstract__icontains'] = self.data.get('abstract')

        if select_ids is not None and len(select_ids) > 0:
            base_query_dict['application_chat.id__in'] = select_ids
        base_condition = Q(**base_query_dict)
        min_star_query = None
        min_trample_query = None
        if 'min_star' in self.data and self.data.get('min_star') is not None:
            min_star_query = Q(star_num__gte=self.data.get('min_star'))
        if 'min_trample' in self.data and self.data.get('min_trample') is not None:
            min_trample_query = Q(trample_num__gte=self.data.get('min_trample'))
        if min_star_query is not None and min_trample_query is not None:
            if self.data.get(
                    'comparer') is not None and self.data.get('comparer') == 'or':
                condition = base_condition & (min_star_query | min_trample_query)
            else:
                condition = base_condition & (min_star_query & min_trample_query)
        elif min_star_query is not None:
            condition = base_condition & min_star_query
        elif min_trample_query is not None:
            condition = base_condition & min_trample_query
        else:
            condition = base_condition

        return {
            'default_queryset': query_set.filter(condition).order_by("-application_chat.update_time")
        }

    def list(self, with_valid=True):
        if with_valid:
            self.is_valid(raise_exception=True)
        return native_search(self.get_query_set(), select_string=get_file_content(
            os.path.join(PROJECT_DIR, "apps", "application", 'sql',
                         ('list_application_chat_ee.sql' if ['PE', 'EE'].__contains__(
                             edition) else 'list_application_chat.sql'))),
                             with_table_name=False)

    @staticmethod
    def paragraph_list_to_string(paragraph_list):
        return "\n**********\n".join(
            [f"{paragraph.get('title')}:\n{paragraph.get('content')}" for paragraph in
             paragraph_list] if paragraph_list is not None else '')

    @staticmethod
    def to_row(row: Dict):
        details = row.get('details') or {}
        padding_problem_text = ' '.join(node.get("answer", "") for key, node in details.items() if
                                        node.get("type") == 'question-node')
        search_dataset_node_list = [(key, node) for key, node in details.items() if
                                    node.get("type") == 'search-dataset-node' or node.get(
                                        "step_type") == 'search_step']
        reference_paragraph_len = '\n'.join([str(len(node.get('paragraph_list',
                                                              []))) if key == 'search_step' else node.get(
            'name') + ':' + str(
            len(node.get('paragraph_list', [])) if node.get('paragraph_list', []) is not None else '0') for
                                             key, node in search_dataset_node_list])
        reference_paragraph = '\n----------\n'.join(
            [ApplicationChatQuerySerializers.paragraph_list_to_string(node.get('paragraph_list',
                                                                               [])) if key == 'search_step' else node.get(
                'name') + ':\n' + ApplicationChatQuerySerializers.paragraph_list_to_string(node.get('paragraph_list',
                                                                                                    [])) for
             key, node in search_dataset_node_list])
        improve_paragraph_list = row.get('improve_paragraph_list') or []
        vote_status_map = {'-1': '未投票', '0': '赞同', '1': '反对'}
        return [str(row.get('chat_id')), row.get('abstract'), row.get('problem_text'), padding_problem_text,
                row.get('answer_text'), vote_status_map.get(row.get('vote_status')), reference_paragraph_len,
                reference_paragraph,
                "\n".join([
                    f"{improve_paragraph_list[index].get('title')}\n{improve_paragraph_list[index].get('content')}"
                    for index in range(len(improve_paragraph_list))]),
                row.get('asker').get('username'),
                (row.get('message_tokens') or 0) + (row.get('answer_tokens') or 0), row.get('run_time'),
                str(row.get('create_time').astimezone(pytz.timezone(TIME_ZONE)).strftime('%Y-%m-%d %H:%M:%S')
                    if row.get('create_time') is not None else None)]

    def export(self, data, with_valid=True):
        if with_valid:
            self.is_valid(raise_exception=True)
        ApplicationChatRecordExportRequest(data=data).is_valid(raise_exception=True)

        data_list = native_search(self.get_query_set(data.get('select_ids')),
                                  select_string=get_file_content(
                                      os.path.join(PROJECT_DIR, "apps", "application", 'sql',
                                                   ('export_application_chat_ee.sql' if ['PE', 'EE'].__contains__(
                                                       edition) else 'export_application_chat.sql'))),
                                  with_table_name=False)

        batch_size = 500

        def stream_response():
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = 'Sheet1'

            headers = [gettext('Conversation ID'), gettext('summary'), gettext('User Questions'),
                       gettext('Problem after optimization'),
                       gettext('answer'), gettext('User feedback'),
                       gettext('Reference segment number'),
                       gettext('Section title + content'),
                       gettext('Annotation'), gettext('USER'), gettext('Consuming tokens'),
                       gettext('Time consumed (s)'),
                       gettext('Question Time')]
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header

            for i in range(0, len(data_list), batch_size):
                batch_data = data_list[i:i + batch_size]

                for row_idx, row in enumerate(batch_data, start=i + 2):
                    for col_idx, value in enumerate(self.to_row(row), 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        if isinstance(value, str):
                            value = re.sub(ILLEGAL_CHARACTERS_RE, '', value)
                        if isinstance(value, datetime.datetime):
                            eastern = pytz.timezone(TIME_ZONE)
                            c = datetime.timezone(eastern._utcoffset)
                            value = value.astimezone(c)
                        cell.value = value

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            yield output.getvalue()
            output.close()
            workbook.close()

        response = StreamingHttpResponse(stream_response(),
                                         content_type='application/vnd.open.xmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="data.xlsx"'
        return response

    def page(self, current_page: int, page_size: int, with_valid=True):
        if with_valid:
            self.is_valid(raise_exception=True)
        return native_page_search(current_page, page_size, self.get_query_set(), select_string=get_file_content(
            os.path.join(PROJECT_DIR, "apps", "application", 'sql',
                         ('list_application_chat_ee.sql' if ['PE', 'EE'].__contains__(
                             edition) else 'list_application_chat.sql'))),
                                  with_table_name=False)


class ChatCountSerializer(serializers.Serializer):
    chat_id = serializers.UUIDField(required=True, label=_("Conversation ID"))

    def get_query_set(self):
        return QuerySet(ChatRecord).filter(chat_id=self.data.get('chat_id'))

    def update_chat(self):
        self.is_valid(raise_exception=True)
        count_chat_record = native_search(self.get_query_set(), get_file_content(
            os.path.join(PROJECT_DIR, "apps", "application", 'sql', 'count_chat_record.sql')), with_search_one=True)
        QuerySet(Chat).filter(id=self.data.get('chat_id')).update(star_num=count_chat_record.get('star_num', 0) or 0,
                                                                  trample_num=count_chat_record.get('trample_num',
                                                                                                    0) or 0,
                                                                  chat_record_count=count_chat_record.get(
                                                                      'chat_record_count', 0) or 0,
                                                                  mark_sum=count_chat_record.get('mark_sum', 0) or 0)
        return True
