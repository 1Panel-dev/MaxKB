# coding=utf-8
"""
    @project: maxkb
    @Author：虎
    @file： chat_serializers.py
    @date：2023/11/14 9:59
    @desc:
"""
import datetime
import os
import re
import uuid
from functools import reduce
from io import BytesIO
from typing import Dict
import pytz
import openpyxl
from django.core import validators
from django.core.cache import caches
from django.db import transaction, models
from django.db.models import QuerySet, Q
from django.http import StreamingHttpResponse
from django.utils.translation import gettext_lazy as _, gettext
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from rest_framework import serializers
from rest_framework.utils.formatting import lazy_format

from application.flow.workflow_manage import Flow
from application.models import Chat, Application, ApplicationDatasetMapping, VoteChoices, ChatRecord, WorkFlowVersion, \
    ApplicationTypeChoices
from application.models.api_key_model import ApplicationAccessToken
from application.serializers.application_serializers import ModelDatasetAssociation, DatasetSettingSerializer, \
    ModelSettingSerializer
from application.serializers.chat_message_serializers import ChatInfo
from common.constants.permission_constants import RoleConstants
from common.db.search import native_search, native_page_search, page_search, get_dynamics_model
from common.exception.app_exception import AppApiException
from common.util.common import post
from common.util.field_message import ErrMessage
from common.util.file_util import get_file_content
from common.util.lock import try_lock, un_lock
from dataset.models import Document, Problem, Paragraph, ProblemParagraphMapping
from dataset.serializers.common_serializers import get_embedding_model_id_by_dataset_id, update_document_char_length
from dataset.serializers.paragraph_serializers import ParagraphSerializers
from embedding.task import embedding_by_paragraph, embedding_by_paragraph_list
from setting.models import Model
from setting.models_provider import get_model_credential
from smartdoc.conf import PROJECT_DIR
from smartdoc.settings import TIME_ZONE

chat_cache = caches['chat_cache']


class WorkFlowSerializers(serializers.Serializer):
    nodes = serializers.ListSerializer(child=serializers.DictField(), error_messages=ErrMessage.uuid(_("node")))
    edges = serializers.ListSerializer(child=serializers.DictField(), error_messages=ErrMessage.uuid(_("Connection")))


def valid_model_params_setting(model_id, model_params_setting):
    if model_id is None:
        return
    model = QuerySet(Model).filter(id=model_id).first()
    credential = get_model_credential(model.provider, model.model_type, model.model_name)
    model_params_setting_form = credential.get_model_params_setting_form(model.model_name)
    if model_params_setting is None or len(model_params_setting.keys()) == 0:
        model_params_setting = model_params_setting_form.get_default_form_data()
    credential.get_model_params_setting_form(model.model_name).valid_form(model_params_setting)


class ReAbstractInstanceSerializers(serializers.Serializer):
    abstract = serializers.CharField(required=True, error_messages=ErrMessage.char(_("abstract")))


class ChatSerializers(serializers.Serializer):
    class Operate(serializers.Serializer):
        chat_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Conversation ID")))
        application_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Application ID")))

        def logic_delete(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            QuerySet(Chat).filter(id=self.data.get('chat_id'), application_id=self.data.get('application_id')).update(
                is_deleted=True)
            return True

        def re_abstract(self, instance, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                ReAbstractInstanceSerializers(data=instance).is_valid(raise_exception=True)

            QuerySet(Chat).filter(id=self.data.get('chat_id'), application_id=self.data.get('application_id')).update(
                abstract=instance.get('abstract'))
            return True

        def delete(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            QuerySet(Chat).filter(id=self.data.get('chat_id'), application_id=self.data.get('application_id')).delete()
            return True

    class ClientChatHistory(serializers.Serializer):
        application_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Application ID")))
        client_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Client id")))

        def page(self, current_page: int, page_size: int, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            queryset = QuerySet(Chat).filter(client_id=self.data.get('client_id'),
                                             application_id=self.data.get('application_id'),
                                             is_deleted=False)
            queryset = queryset.order_by('-create_time')
            return page_search(current_page, page_size, queryset, lambda row: ChatSerializerModel(row).data)

    class Query(serializers.Serializer):
        abstract = serializers.CharField(required=False, error_messages=ErrMessage.char(_("summary")))
        start_time = serializers.DateField(format='%Y-%m-%d', error_messages=ErrMessage.date(_("Start time")))
        end_time = serializers.DateField(format='%Y-%m-%d', error_messages=ErrMessage.date(_("End time")))
        user_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("User ID")))
        application_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Application ID")))
        min_star = serializers.IntegerField(required=False, min_value=0,
                                            error_messages=ErrMessage.integer(_("Minimum number of likes")))
        min_trample = serializers.IntegerField(required=False, min_value=0,
                                               error_messages=ErrMessage.integer(_("Minimum number of clicks")))
        comparer = serializers.CharField(required=False, error_messages=ErrMessage.char(_("Comparator")), validators=[
            validators.RegexValidator(regex=re.compile("^and|or$"),
                                      message=_("Only supports and|or"), code=500)
        ])

        def get_end_time(self):
            return datetime.datetime.combine(
                datetime.datetime.strptime(self.data.get('end_time'), '%Y-%m-%d'),
                datetime.datetime.max.time())

        def get_start_time(self):
            return self.data.get('start_time')

        def get_query_set(self, select_ids=None):
            end_time = self.get_end_time()
            start_time = self.get_start_time()
            query_set = QuerySet(model=get_dynamics_model(
                {'application_chat.application_id': models.CharField(),
                 'application_chat.abstract': models.CharField(),
                 "star_num": models.IntegerField(),
                 'trample_num': models.IntegerField(),
                 'comparer': models.CharField(),
                 'application_chat.update_time': models.DateTimeField(),
                 'application_chat.id': models.UUIDField(), }))

            base_query_dict = {'application_chat.application_id': self.data.get("application_id"),
                               'application_chat.update_time__gte': start_time,
                               'application_chat.update_time__lte': end_time,
                               }
            if 'abstract' in self.data and self.data.get('abstract') is not None:
                base_query_dict['application_chat.abstract__icontains'] = self.data.get('abstract')

            if select_ids is not None and len(select_ids) > 0:
                base_query_dict['application_chat.id__in'] = select_ids
            base_condition = Q(**base_query_dict)
            min_star_query = None
            min_trample_query = None
            if 'min_star' in self.data and self.data.get('min_star') is not None:
                min_star_query = Q(star_num__gte=self.data.get('min_star'))
            if 'min_trample' in self.data and self.data.get('min_trample') is not None:
                min_trample_query = Q(trample_num__gte=self.data.get('min_trample'))
            if min_star_query is not None and min_trample_query is not None:
                if self.data.get(
                        'comparer') is not None and self.data.get('comparer') == 'or':
                    condition = base_condition & (min_star_query | min_trample_query)
                else:
                    condition = base_condition & (min_star_query & min_trample_query)
            elif min_star_query is not None:
                condition = base_condition & min_star_query
            elif min_trample_query is not None:
                condition = base_condition & min_trample_query
            else:
                condition = base_condition
            inner_queryset = QuerySet(Chat).filter(application_id=self.data.get("application_id"))
            if 'abstract' in self.data and self.data.get('abstract') is not None:
                inner_queryset = inner_queryset.filter(abstract__icontains=self.data.get('abstract'))

            return {
                'inner_queryset': inner_queryset,
                'default_queryset': query_set.filter(condition).order_by("-application_chat.update_time")
            }

        def list(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            return native_search(self.get_query_set(), select_string=get_file_content(
                os.path.join(PROJECT_DIR, "apps", "application", 'sql', 'list_application_chat.sql')),
                                 with_table_name=False)

        @staticmethod
        def paragraph_list_to_string(paragraph_list):
            return "\n**********\n".join(
                [f"{paragraph.get('title')}:\n{paragraph.get('content')}" for paragraph in
                 paragraph_list] if paragraph_list is not None else '')

        @staticmethod
        def to_row(row: Dict):
            details = row.get('details')
            padding_problem_text = ' '.join(node.get("answer", "") for key, node in details.items() if
                                            node.get("type") == 'question-node')
            search_dataset_node_list = [(key, node) for key, node in details.items() if
                                        node.get("type") == 'search-dataset-node' or node.get(
                                            "step_type") == 'search_step']
            reference_paragraph_len = '\n'.join([str(len(node.get('paragraph_list',
                                                                  []))) if key == 'search_step' else node.get(
                'name') + ':' + str(
                len(node.get('paragraph_list', [])) if node.get('paragraph_list', []) is not None else '0') for
                                                 key, node in search_dataset_node_list])
            reference_paragraph = '\n----------\n'.join(
                [ChatSerializers.Query.paragraph_list_to_string(node.get('paragraph_list',
                                                                         [])) if key == 'search_step' else node.get(
                    'name') + ':\n' + ChatSerializers.Query.paragraph_list_to_string(node.get('paragraph_list',
                                                                                              [])) for
                 key, node in search_dataset_node_list])
            improve_paragraph_list = row.get('improve_paragraph_list')
            vote_status_map = {'-1': '未投票', '0': '赞同', '1': '反对'}
            return [str(row.get('chat_id')), row.get('abstract'), row.get('problem_text'), padding_problem_text,
                    row.get('answer_text'), vote_status_map.get(row.get('vote_status')), reference_paragraph_len,
                    reference_paragraph,
                    "\n".join([
                        f"{improve_paragraph_list[index].get('title')}\n{improve_paragraph_list[index].get('content')}"
                        for index in range(len(improve_paragraph_list))]),
                    row.get('asker').get('user_name'),
                    row.get('message_tokens') + row.get('answer_tokens'), row.get('run_time'),
                    str(row.get('create_time').astimezone(pytz.timezone(TIME_ZONE)).strftime('%Y-%m-%d %H:%M:%S')
                        )]

        def export(self, data, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)

            data_list = native_search(self.get_query_set(data.get('select_ids')),
                                      select_string=get_file_content(
                                          os.path.join(PROJECT_DIR, "apps", "application", 'sql',
                                                       'export_application_chat.sql')),
                                      with_table_name=False)

            batch_size = 500

            def stream_response():
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = 'Sheet1'

                headers = [gettext('Conversation ID'), gettext('summary'), gettext('User Questions'),
                           gettext('Problem after optimization'),
                           gettext('answer'), gettext('User feedback'),
                           gettext('Reference segment number'),
                           gettext('Section title + content'),
                           gettext('Annotation'), gettext('USER'), gettext('Consuming tokens'),
                           gettext('Time consumed (s)'),
                           gettext('Question Time')]
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=1, column=col_idx)
                    cell.value = header

                for i in range(0, len(data_list), batch_size):
                    batch_data = data_list[i:i + batch_size]

                    for row_idx, row in enumerate(batch_data, start=i + 2):
                        for col_idx, value in enumerate(self.to_row(row), 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if isinstance(value, str):
                                value = re.sub(ILLEGAL_CHARACTERS_RE, '', value)
                            if isinstance(value, datetime.datetime):
                                eastern = pytz.timezone(TIME_ZONE)
                                c = datetime.timezone(eastern._utcoffset)
                                value = value.astimezone(c)
                            cell.value = value

                output = BytesIO()
                workbook.save(output)
                output.seek(0)
                yield output.getvalue()
                output.close()
                workbook.close()

            response = StreamingHttpResponse(stream_response(),
                                             content_type='application/vnd.open.xmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="data.xlsx"'
            return response

        def page(self, current_page: int, page_size: int, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            return native_page_search(current_page, page_size, self.get_query_set(), select_string=get_file_content(
                os.path.join(PROJECT_DIR, "apps", "application", 'sql', 'list_application_chat.sql')),
                                      with_table_name=False)

    class OpenChat(serializers.Serializer):
        user_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("User ID")))

        application_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Application ID")))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            user_id = self.data.get('user_id')
            application_id = self.data.get('application_id')
            if not QuerySet(Application).filter(id=application_id, user_id=user_id).exists():
                raise AppApiException(500, gettext('Application does not exist'))

        def open(self):
            self.is_valid(raise_exception=True)
            application_id = self.data.get('application_id')
            application = QuerySet(Application).get(id=application_id)
            if application.type == ApplicationTypeChoices.SIMPLE:
                return self.open_simple(application)
            else:
                return self.open_work_flow(application)

        def open_work_flow(self, application):
            self.is_valid(raise_exception=True)
            application_id = self.data.get('application_id')
            chat_id = str(uuid.uuid1())
            work_flow_version = QuerySet(WorkFlowVersion).filter(application_id=application_id).order_by(
                '-create_time')[0:1].first()
            if work_flow_version is None:
                raise AppApiException(500,
                                      gettext(
                                          "The application has not been published. Please use it after publishing."))
            chat_cache.set(chat_id,
                           ChatInfo(chat_id, [],
                                    [],
                                    application, work_flow_version), timeout=60 * 30)
            return chat_id

        def open_simple(self, application):
            application_id = self.data.get('application_id')
            dataset_id_list = [str(row.dataset_id) for row in
                               QuerySet(ApplicationDatasetMapping).filter(
                                   application_id=application_id)]
            chat_id = str(uuid.uuid1())
            chat_cache.set(chat_id,
                           ChatInfo(chat_id, dataset_id_list,
                                    [str(document.id) for document in
                                     QuerySet(Document).filter(
                                         dataset_id__in=dataset_id_list,
                                         is_active=False)],
                                    application), timeout=60 * 30)
            return chat_id

    class OpenWorkFlowChat(serializers.Serializer):
        work_flow = WorkFlowSerializers(error_messages=ErrMessage.uuid(_("Workflow")))
        user_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("User ID")))

        def open(self):
            self.is_valid(raise_exception=True)
            work_flow = self.data.get('work_flow')
            Flow.new_instance(work_flow).is_valid()
            chat_id = str(uuid.uuid1())
            application = Application(id=None, dialogue_number=3, model=None,
                                      dataset_setting={},
                                      model_setting={},
                                      problem_optimization=None,
                                      type=ApplicationTypeChoices.WORK_FLOW,
                                      user_id=self.data.get('user_id')
                                      )
            work_flow_version = WorkFlowVersion(work_flow=work_flow)
            chat_cache.set(chat_id,
                           ChatInfo(chat_id, [],
                                    [],
                                    application, work_flow_version), timeout=60 * 30)
            return chat_id

    class OpenTempChat(serializers.Serializer):
        user_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("User ID")))

        id = serializers.UUIDField(required=False, allow_null=True,
                                   error_messages=ErrMessage.uuid(_("Application ID")))
        model_id = serializers.CharField(required=False, allow_null=True, allow_blank=True,
                                         error_messages=ErrMessage.uuid(_("Model id")))

        multiple_rounds_dialogue = serializers.BooleanField(required=True,
                                                            error_messages=ErrMessage.boolean(
                                                                _("Multi-round conversation")))

        dataset_id_list = serializers.ListSerializer(required=False, child=serializers.UUIDField(required=True),
                                                     error_messages=ErrMessage.list(_("Related Datasets")))
        # 数据集相关设置
        dataset_setting = DatasetSettingSerializer(required=True)
        # 模型相关设置
        model_setting = ModelSettingSerializer(required=True)
        # 问题补全
        problem_optimization = serializers.BooleanField(required=True,
                                                        error_messages=ErrMessage.boolean(_("Question completion")))
        # 模型相关设置
        model_params_setting = serializers.JSONField(required=False,
                                                     error_messages=ErrMessage.dict(_("Model parameter settings")))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            user_id = self.get_user_id()
            ModelDatasetAssociation(
                data={'user_id': user_id, 'model_id': self.data.get('model_id'),
                      'dataset_id_list': self.data.get('dataset_id_list')}).is_valid()
            return user_id

        def get_user_id(self):
            if 'id' in self.data and self.data.get('id') is not None:
                application = QuerySet(Application).filter(id=self.data.get('id')).first()
                if application is None:
                    raise AppApiException(500, gettext("Application does not exist"))
                return application.user_id
            return self.data.get('user_id')

        def open(self):
            user_id = self.is_valid(raise_exception=True)
            chat_id = str(uuid.uuid1())
            model_id = self.data.get('model_id')
            dataset_id_list = self.data.get('dataset_id_list')
            dialogue_number = 3 if self.data.get('multiple_rounds_dialogue', False) else 0
            application = Application(id=None, dialogue_number=dialogue_number, model_id=model_id,
                                      dataset_setting=self.data.get('dataset_setting'),
                                      model_setting=self.data.get('model_setting'),
                                      problem_optimization=self.data.get('problem_optimization'),
                                      model_params_setting=self.data.get('model_params_setting'),
                                      user_id=user_id)
            chat_cache.set(chat_id,
                           ChatInfo(chat_id, dataset_id_list,
                                    [str(document.id) for document in
                                     QuerySet(Document).filter(
                                         dataset_id__in=dataset_id_list,
                                         is_active=False)],
                                    application), timeout=60 * 30)
            return chat_id


class ChatRecordSerializerModel(serializers.ModelSerializer):
    class Meta:
        model = ChatRecord
        fields = ['id', 'chat_id', 'vote_status', 'problem_text', 'answer_text',
                  'message_tokens', 'answer_tokens', 'const', 'improve_paragraph_id_list', 'run_time', 'index',
                  'answer_text_list',
                  'create_time', 'update_time']


class ChatSerializerModel(serializers.ModelSerializer):
    class Meta:
        model = Chat
        fields = ['id', 'application_id', 'abstract', 'client_id']


class ChatRecordSerializer(serializers.Serializer):
    class Operate(serializers.Serializer):
        chat_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Conversation ID")))
        application_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Application ID")))
        chat_record_id = serializers.UUIDField(required=True,
                                               error_messages=ErrMessage.uuid(_("Conversation record id")))

        def is_valid(self, *, current_role=None, raise_exception=False):
            super().is_valid(raise_exception=True)
            application_access_token = QuerySet(ApplicationAccessToken).filter(
                application_id=self.data.get('application_id')).first()
            if application_access_token is None:
                raise AppApiException(500, gettext('Application authentication information does not exist'))
            if not application_access_token.show_source and current_role == RoleConstants.APPLICATION_ACCESS_TOKEN.value:
                raise AppApiException(500, gettext('Displaying knowledge sources is not enabled'))

        def get_chat_record(self):
            chat_record_id = self.data.get('chat_record_id')
            chat_id = self.data.get('chat_id')
            chat_info: ChatInfo = chat_cache.get(chat_id)
            if chat_info is not None:
                chat_record_list = [chat_record for chat_record in chat_info.chat_record_list if
                                    str(chat_record.id) == str(chat_record_id)]
                if chat_record_list is not None and len(chat_record_list):
                    return chat_record_list[-1]
            return QuerySet(ChatRecord).filter(id=chat_record_id, chat_id=chat_id).first()

        def one(self, current_role: RoleConstants, with_valid=True):
            if with_valid:
                self.is_valid(current_role=current_role, raise_exception=True)
            chat_record = self.get_chat_record()
            if chat_record is None:
                raise AppApiException(500, gettext("Conversation does not exist"))
            return ChatRecordSerializer.Query.reset_chat_record(chat_record)

    class Query(serializers.Serializer):
        application_id = serializers.UUIDField(required=True)
        chat_id = serializers.UUIDField(required=True)
        order_asc = serializers.BooleanField(required=False, allow_null=True)

        def list(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            QuerySet(ChatRecord).filter(chat_id=self.data.get('chat_id'))
            order_by = 'create_time' if self.data.get('order_asc') is None or self.data.get(
                'order_asc') else '-create_time'
            return [ChatRecordSerializerModel(chat_record).data for chat_record in
                    QuerySet(ChatRecord).filter(chat_id=self.data.get('chat_id')).order_by(order_by)]

        @staticmethod
        def reset_chat_record(chat_record):
            dataset_list = []
            paragraph_list = []

            if 'search_step' in chat_record.details and chat_record.details.get('search_step').get(
                    'paragraph_list') is not None:
                paragraph_list = chat_record.details.get('search_step').get(
                    'paragraph_list')
                dataset_list = [{'id': dataset_id, 'name': name} for dataset_id, name in reduce(lambda x, y: {**x, **y},
                                                                                                [{row.get(
                                                                                                    'dataset_id'): row.get(
                                                                                                    "dataset_name")} for
                                                                                                    row in
                                                                                                    paragraph_list],
                                                                                                {}).items()]
            if len(chat_record.improve_paragraph_id_list) > 0:
                paragraph_model_list = QuerySet(Paragraph).filter(id__in=chat_record.improve_paragraph_id_list)
                if len(paragraph_model_list) < len(chat_record.improve_paragraph_id_list):
                    paragraph_model_id_list = [str(p.id) for p in paragraph_model_list]
                    chat_record.improve_paragraph_id_list = list(
                        filter(lambda p_id: paragraph_model_id_list.__contains__(p_id),
                               chat_record.improve_paragraph_id_list))
                    chat_record.save()

            return {
                **ChatRecordSerializerModel(chat_record).data,
                'padding_problem_text': chat_record.details.get('problem_padding').get(
                    'padding_problem_text') if 'problem_padding' in chat_record.details else None,
                'dataset_list': dataset_list,
                'paragraph_list': paragraph_list,
                'execution_details': [chat_record.details[key] for key in chat_record.details]
            }

        def page(self, current_page: int, page_size: int, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            order_by = '-create_time' if self.data.get('order_asc') is None or self.data.get(
                'order_asc') else 'create_time'
            page = page_search(current_page, page_size,
                               QuerySet(ChatRecord).filter(chat_id=self.data.get('chat_id')).order_by(order_by),
                               post_records_handler=lambda chat_record: self.reset_chat_record(chat_record))
            return page

    class Vote(serializers.Serializer):
        chat_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Conversation ID")))

        chat_record_id = serializers.UUIDField(required=True,
                                               error_messages=ErrMessage.uuid(_("Conversation record id")))

        vote_status = serializers.ChoiceField(choices=VoteChoices.choices,
                                              error_messages=ErrMessage.uuid(_("Bidding Status")))

        @transaction.atomic
        def vote(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            if not try_lock(self.data.get('chat_record_id')):
                raise AppApiException(500,
                                      gettext(
                                          "Voting on the current session minutes, please do not send repeated requests"))
            try:
                chat_record_details_model = QuerySet(ChatRecord).get(id=self.data.get('chat_record_id'),
                                                                     chat_id=self.data.get('chat_id'))
                if chat_record_details_model is None:
                    raise AppApiException(500, gettext("Non-existent conversation chat_record_id"))
                vote_status = self.data.get("vote_status")
                if chat_record_details_model.vote_status == VoteChoices.UN_VOTE:
                    if vote_status == VoteChoices.STAR:
                        # 点赞
                        chat_record_details_model.vote_status = VoteChoices.STAR

                    if vote_status == VoteChoices.TRAMPLE:
                        # 点踩
                        chat_record_details_model.vote_status = VoteChoices.TRAMPLE
                    chat_record_details_model.save()
                else:
                    if vote_status == VoteChoices.UN_VOTE:
                        # 取消点赞
                        chat_record_details_model.vote_status = VoteChoices.UN_VOTE
                        chat_record_details_model.save()
                    else:
                        raise AppApiException(500, gettext("Already voted, please cancel first and then vote again"))
            finally:
                un_lock(self.data.get('chat_record_id'))
            return True

    class ImproveSerializer(serializers.Serializer):
        title = serializers.CharField(required=False, max_length=256, allow_null=True, allow_blank=True,
                                      error_messages=ErrMessage.char(_("Section title")))
        content = serializers.CharField(required=True, error_messages=ErrMessage.char(_("Paragraph content")))

        problem_text = serializers.CharField(required=False, max_length=256, allow_null=True, allow_blank=True,
                                             error_messages=ErrMessage.char(_("question")))

    class ParagraphModel(serializers.ModelSerializer):
        class Meta:
            model = Paragraph
            fields = "__all__"

    class ChatRecordImprove(serializers.Serializer):
        chat_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Conversation ID")))

        chat_record_id = serializers.UUIDField(required=True,
                                               error_messages=ErrMessage.uuid(_("Conversation record id")))

        def get(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            chat_record_id = self.data.get('chat_record_id')
            chat_id = self.data.get('chat_id')
            chat_record = QuerySet(ChatRecord).filter(id=chat_record_id, chat_id=chat_id).first()
            if chat_record is None:
                raise AppApiException(500, gettext('Conversation record does not exist'))
            if chat_record.improve_paragraph_id_list is None or len(chat_record.improve_paragraph_id_list) == 0:
                return []

            paragraph_model_list = QuerySet(Paragraph).filter(id__in=chat_record.improve_paragraph_id_list)
            if len(paragraph_model_list) < len(chat_record.improve_paragraph_id_list):
                paragraph_model_id_list = [str(p.id) for p in paragraph_model_list]
                chat_record.improve_paragraph_id_list = list(
                    filter(lambda p_id: paragraph_model_id_list.__contains__(p_id),
                           chat_record.improve_paragraph_id_list))
                chat_record.save()
            return [ChatRecordSerializer.ParagraphModel(p).data for p in paragraph_model_list]

    class Improve(serializers.Serializer):
        chat_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Conversation ID")))

        chat_record_id = serializers.UUIDField(required=True,
                                               error_messages=ErrMessage.uuid(_("Conversation record id")))

        dataset_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Knowledge base id")))

        document_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Document id")))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            if not QuerySet(Document).filter(id=self.data.get('document_id'),
                                             dataset_id=self.data.get('dataset_id')).exists():
                raise AppApiException(500, gettext("The document id is incorrect"))

        @staticmethod
        def post_embedding_paragraph(chat_record, paragraph_id, dataset_id):
            model_id = get_embedding_model_id_by_dataset_id(dataset_id)
            # 发送向量化事件
            embedding_by_paragraph(paragraph_id, model_id)
            return chat_record

        @post(post_function=post_embedding_paragraph)
        @transaction.atomic
        def improve(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            ChatRecordSerializer.ImproveSerializer(data=instance).is_valid(raise_exception=True)
            chat_record_id = self.data.get('chat_record_id')
            chat_id = self.data.get('chat_id')
            chat_record = QuerySet(ChatRecord).filter(id=chat_record_id, chat_id=chat_id).first()
            if chat_record is None:
                raise AppApiException(500, gettext('Conversation record does not exist'))

            document_id = self.data.get("document_id")
            dataset_id = self.data.get("dataset_id")
            paragraph = Paragraph(id=uuid.uuid1(),
                                  document_id=document_id,
                                  content=instance.get("content"),
                                  dataset_id=dataset_id,
                                  title=instance.get("title") if 'title' in instance else '')
            problem_text = instance.get('problem_text') if instance.get(
                'problem_text') is not None else chat_record.problem_text
            problem, _ = Problem.objects.get_or_create(content=problem_text, dataset_id=dataset_id)
            problem_paragraph_mapping = ProblemParagraphMapping(id=uuid.uuid1(), dataset_id=dataset_id,
                                                                document_id=document_id,
                                                                problem_id=problem.id,
                                                                paragraph_id=paragraph.id)
            # 插入段落
            paragraph.save()
            # 插入关联问题
            problem_paragraph_mapping.save()
            chat_record.improve_paragraph_id_list.append(paragraph.id)
            update_document_char_length(document_id)
            # 添加标注
            chat_record.save()
            return ChatRecordSerializerModel(chat_record).data, paragraph.id, dataset_id

        class Operate(serializers.Serializer):
            chat_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Conversation ID")))

            chat_record_id = serializers.UUIDField(required=True,
                                                   error_messages=ErrMessage.uuid(_("Conversation record id")))

            dataset_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Knowledge base id")))

            document_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Document id")))

            paragraph_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Paragraph id")))

            def delete(self, with_valid=True):
                if with_valid:
                    self.is_valid(raise_exception=True)

                chat_record_id = self.data.get('chat_record_id')
                chat_id = self.data.get('chat_id')
                dataset_id = self.data.get('dataset_id')
                document_id = self.data.get('document_id')
                paragraph_id = self.data.get('paragraph_id')
                chat_record = QuerySet(ChatRecord).filter(id=chat_record_id, chat_id=chat_id).first()
                if chat_record is None:
                    raise AppApiException(500, gettext('Conversation record does not exist'))
                if not chat_record.improve_paragraph_id_list.__contains__(uuid.UUID(paragraph_id)):
                    message = lazy_format(
                        _('The paragraph id is wrong. The current conversation record does not exist. [{paragraph_id}] paragraph id'),
                        paragraph_id=paragraph_id)
                    raise AppApiException(500, message)
                chat_record.improve_paragraph_id_list = [row for row in chat_record.improve_paragraph_id_list if
                                                         str(row) != paragraph_id]
                chat_record.save()
                o = ParagraphSerializers.Operate(
                    data={"dataset_id": dataset_id, 'document_id': document_id, "paragraph_id": paragraph_id})
                o.is_valid(raise_exception=True)
                return o.delete()

    class PostImprove(serializers.Serializer):
        dataset_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Knowledge base id")))
        document_id = serializers.UUIDField(required=True, error_messages=ErrMessage.uuid(_("Document id")))
        chat_ids = serializers.ListSerializer(child=serializers.UUIDField(), required=True,
                                              error_messages=ErrMessage.list(_("Conversation ID")))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            if not Document.objects.filter(id=self.data['document_id'], dataset_id=self.data['dataset_id']).exists():
                raise AppApiException(500, gettext("The document id is incorrect"))

        @staticmethod
        def post_embedding_paragraph(paragraph_ids, dataset_id):
            model_id = get_embedding_model_id_by_dataset_id(dataset_id)
            embedding_by_paragraph_list(paragraph_ids, model_id)

        @post(post_function=post_embedding_paragraph)
        @transaction.atomic
        def post_improve(self, instance: Dict):
            ChatRecordSerializer.PostImprove(data=instance).is_valid(raise_exception=True)

            chat_ids = instance['chat_ids']
            document_id = instance['document_id']
            dataset_id = instance['dataset_id']

            # 获取所有聊天记录
            chat_record_list = list(ChatRecord.objects.filter(chat_id__in=chat_ids))
            if len(chat_record_list) < len(chat_ids):
                raise AppApiException(500, gettext("Conversation records that do not exist"))

            # 批量创建段落和问题映射
            paragraphs = []
            paragraph_ids = []
            problem_paragraph_mappings = []
            for chat_record in chat_record_list:
                paragraph = Paragraph(
                    id=uuid.uuid1(),
                    document_id=document_id,
                    content=chat_record.answer_text,
                    dataset_id=dataset_id,
                    title=chat_record.problem_text
                )
                problem, _ = Problem.objects.get_or_create(content=chat_record.problem_text, dataset_id=dataset_id)
                problem_paragraph_mapping = ProblemParagraphMapping(
                    id=uuid.uuid1(),
                    dataset_id=dataset_id,
                    document_id=document_id,
                    problem_id=problem.id,
                    paragraph_id=paragraph.id
                )
                paragraphs.append(paragraph)
                paragraph_ids.append(paragraph.id)
                problem_paragraph_mappings.append(problem_paragraph_mapping)
                chat_record.improve_paragraph_id_list.append(paragraph.id)

            # 批量保存段落和问题映射
            Paragraph.objects.bulk_create(paragraphs)
            ProblemParagraphMapping.objects.bulk_create(problem_paragraph_mappings)

            # 批量保存聊天记录
            ChatRecord.objects.bulk_update(chat_record_list, ['improve_paragraph_id_list'])
            update_document_char_length(document_id)

            return paragraph_ids, dataset_id
