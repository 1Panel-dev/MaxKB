import io
import os
import re
import traceback
from functools import reduce
from tempfile import TemporaryDirectory
from typing import Dict, List

import openpyxl
import uuid_utils.compat as uuid
from celery_once import AlreadyQueued
from django.core import validators
from django.db import transaction, models
from django.db.models import QuerySet
from django.db.models.aggregates import Max
from django.db.models.functions import Substr, Reverse
from django.http import HttpResponse
from django.utils.translation import gettext_lazy as _, gettext, get_language, to_locale
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from rest_framework import serializers
from xlwt import Utils

from common.db.search import native_search, get_dynamics_model, native_page_search
from common.event import ListenerManagement
from common.event.common import work_thread_pool
from common.exception.app_exception import AppApiException
from common.handle.impl.qa.csv_parse_qa_handle import CsvParseQAHandle
from common.handle.impl.qa.xls_parse_qa_handle import XlsParseQAHandle
from common.handle.impl.qa.xlsx_parse_qa_handle import XlsxParseQAHandle
from common.handle.impl.qa.zip_parse_qa_handle import ZipParseQAHandle
from common.handle.impl.table.csv_parse_table_handle import CsvParseTableHandle
from common.handle.impl.table.xls_parse_table_handle import XlsParseTableHandle
from common.handle.impl.table.xlsx_parse_table_handle import XlsxParseTableHandle
from common.handle.impl.text.csv_split_handle import CsvSplitHandle
from common.handle.impl.text.doc_split_handle import DocSplitHandle
from common.handle.impl.text.html_split_handle import HTMLSplitHandle
from common.handle.impl.text.pdf_split_handle import PdfSplitHandle
from common.handle.impl.text.text_split_handle import TextSplitHandle
from common.handle.impl.text.xls_split_handle import XlsSplitHandle
from common.handle.impl.text.xlsx_split_handle import XlsxSplitHandle
from common.handle.impl.text.zip_split_handle import ZipSplitHandle
from common.utils.common import post, get_file_content, bulk_create_in_batches, parse_image
from common.utils.fork import Fork
from common.utils.logger import maxkb_logger
from common.utils.split_model import get_split_model, flat_map
from knowledge.models import Knowledge, Paragraph, Problem, Document, KnowledgeType, ProblemParagraphMapping, State, \
    TaskType, File, FileSourceType
from knowledge.serializers.common import ProblemParagraphManage, BatchSerializer, \
    get_embedding_model_id_by_knowledge_id, MetaSerializer, write_image, zip_dir
from knowledge.serializers.paragraph import ParagraphSerializers, ParagraphInstanceSerializer, \
    delete_problems_and_mappings
from knowledge.task.embedding import embedding_by_document, delete_embedding_by_document_list, \
    delete_embedding_by_document, delete_embedding_by_paragraph_ids, embedding_by_document_list, \
    update_embedding_knowledge_id
from knowledge.task.generate import generate_related_by_document_id
from knowledge.task.sync import sync_web_document
from maxkb.const import PROJECT_DIR
from models_provider.models import Model
from oss.serializers.file import FileSerializer

default_split_handle = TextSplitHandle()
split_handles = [
    HTMLSplitHandle(),
    DocSplitHandle(),
    PdfSplitHandle(),
    XlsxSplitHandle(),
    XlsSplitHandle(),
    CsvSplitHandle(),
    ZipSplitHandle(),
    default_split_handle
]

parse_qa_handle_list = [XlsParseQAHandle(), CsvParseQAHandle(), XlsxParseQAHandle(), ZipParseQAHandle()]
parse_table_handle_list = [CsvParseTableHandle(), XlsParseTableHandle(), XlsxParseTableHandle()]


def convert_uuid_to_str(obj):
    if isinstance(obj, dict):
        return {k: convert_uuid_to_str(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_uuid_to_str(i) for i in obj]
    elif isinstance(obj, uuid.UUID):
        return str(obj)
    else:
        return obj


class BatchCancelInstanceSerializer(serializers.Serializer):
    id_list = serializers.ListField(required=True, child=serializers.UUIDField(required=True), label=_('id list'))
    type = serializers.IntegerField(required=True, label=_('task type'))

    def is_valid(self, *, raise_exception=False):
        super().is_valid(raise_exception=True)
        _type = self.data.get('type')
        try:
            TaskType(_type)
        except Exception as e:
            raise AppApiException(500, _('task type not support'))


class DocumentInstanceSerializer(serializers.Serializer):
    name = serializers.CharField(required=True, label=_('document name'), max_length=128, min_length=1,
                                 source=_('document name'))
    paragraphs = ParagraphInstanceSerializer(required=False, many=True, allow_null=True)
    source_file_id = serializers.UUIDField(required=False, allow_null=True, label=_('source file id'))


class CancelInstanceSerializer(serializers.Serializer):
    type = serializers.IntegerField(required=True, label=_('task type'))

    def is_valid(self, *, raise_exception=False):
        super().is_valid(raise_exception=True)
        _type = self.data.get('type')
        try:
            TaskType(_type)
        except Exception as e:
            raise AppApiException(500, _('task type not support'))


class DocumentEditInstanceSerializer(serializers.Serializer):
    meta = serializers.DictField(required=False)
    name = serializers.CharField(required=False, max_length=128, min_length=1, label=_('document name'),
                                 source=_('document name'))
    hit_handling_method = serializers.CharField(required=False, validators=[
        validators.RegexValidator(regex=re.compile("^optimization|directly_return$"),
                                  message=_('The type only supports optimization|directly_return'),
                                  code=500)
    ], label=_('hit handling method'))

    directly_return_similarity = serializers.FloatField(required=False, max_value=2, min_value=0,
                                                        label=_('directly return similarity'))

    is_active = serializers.BooleanField(required=False, label=_('document is active'))

    @staticmethod
    def get_meta_valid_map():
        knowledge_meta_valid_map = {
            KnowledgeType.BASE: MetaSerializer.BaseMeta,
            KnowledgeType.WEB: MetaSerializer.WebMeta
        }
        return knowledge_meta_valid_map

    def is_valid(self, *, document: Document = None):
        super().is_valid(raise_exception=True)
        if 'meta' in self.data and self.data.get('meta') is not None and self.data.get('meta') != {}:
            knowledge_meta_valid_map = self.get_meta_valid_map()
            valid_class = knowledge_meta_valid_map.get(document.type)
            if valid_class is not None:
                valid_class(data=self.data.get('meta')).is_valid(raise_exception=True)


class DocumentSplitRequest(serializers.Serializer):
    file = serializers.ListField(required=True, label=_('file list'))
    limit = serializers.IntegerField(required=False, label=_('limit'))
    patterns = serializers.ListField(
        required=False,
        child=serializers.CharField(required=True, label=_('patterns')),
        label=_('patterns')
    )
    with_filter = serializers.BooleanField(required=False, label=_('Auto Clean'))


class DocumentWebInstanceSerializer(serializers.Serializer):
    source_url_list = serializers.ListField(required=True, label=_('document url list'),
                                            child=serializers.CharField(required=True, label=_('document url list')))
    selector = serializers.CharField(required=False, allow_null=True, allow_blank=True, label=_('selector'))


class DocumentInstanceQASerializer(serializers.Serializer):
    file_list = serializers.ListSerializer(required=True, label=_('file list'),
                                           child=serializers.FileField(required=True, label=_('file')))


class DocumentInstanceTableSerializer(serializers.Serializer):
    file_list = serializers.ListSerializer(required=True, label=_('file list'),
                                           child=serializers.FileField(required=True, label=_('file')))


class DocumentRefreshSerializer(serializers.Serializer):
    state_list = serializers.ListField(required=True, label=_('state list'))


class DocumentBatchRefreshSerializer(serializers.Serializer):
    id_list = serializers.ListField(required=True, label=_('id list'))
    state_list = serializers.ListField(required=True, label=_('state list'))


class DocumentBatchGenerateRelatedSerializer(serializers.Serializer):
    document_id_list = serializers.ListField(required=True, label=_('document id list'))
    model_id = serializers.UUIDField(required=True, label=_('model id'))
    prompt = serializers.CharField(required=True, label=_('prompt'))
    state_list = serializers.ListField(required=True, label=_('state list'))


class DocumentMigrateSerializer(serializers.Serializer):
    document_id_list = serializers.ListField(required=True, label=_('document id list'))


class BatchEditHitHandlingSerializer(serializers.Serializer):
    id_list = serializers.ListField(required=True, child=serializers.UUIDField(required=True), label=_('id list'))
    hit_handling_method = serializers.CharField(required=True, label=_('hit handling method'))
    directly_return_similarity = serializers.FloatField(required=False, max_value=2, min_value=0,
                                                        label=_('directly return similarity'))

    def is_valid(self, *, raise_exception=False):
        super().is_valid(raise_exception=True)
        if self.data.get('hit_handling_method') not in ['optimization', 'directly_return']:
            raise AppApiException(500, _('The type only supports optimization|directly_return'))


class DocumentSerializers(serializers.Serializer):
    class Export(serializers.Serializer):
        type = serializers.CharField(required=True, validators=[
            validators.RegexValidator(regex=re.compile("^csv|excel$"),
                                      message=_('The template type only supports excel|csv'),
                                      code=500)
        ], label=_('type'))

        def export(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            language = get_language()
            if self.data.get('type') == 'csv':
                file = open(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", 'template',
                                 f'csv_template_{to_locale(language)}.csv'),
                    "rb")
                content = file.read()
                file.close()
                return HttpResponse(content, status=200, headers={'Content-Type': 'text/csv',
                                                                  'Content-Disposition': 'attachment; filename="csv_template.csv"'})
            elif self.data.get('type') == 'excel':
                file = open(os.path.join(PROJECT_DIR, "apps", "knowledge", 'template',
                                         f'excel_template_{to_locale(language)}.xlsx'), "rb")
                content = file.read()
                file.close()
                return HttpResponse(content, status=200, headers={'Content-Type': 'application/vnd.ms-excel',
                                                                  'Content-Disposition': 'attachment; filename="excel_template.xlsx"'})
            else:
                return None

        def table_export(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            language = get_language()
            if self.data.get('type') == 'csv':
                file = open(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", 'template',
                                 f'table_template_{to_locale(language)}.csv'),
                    "rb")
                content = file.read()
                file.close()
                return HttpResponse(content, status=200, headers={'Content-Type': 'text/cxv',
                                                                  'Content-Disposition': 'attachment; filename="csv_template.csv"'})
            elif self.data.get('type') == 'excel':
                file = open(os.path.join(PROJECT_DIR, "apps", "knowledge", 'template',
                                         f'table_template_{to_locale(language)}.xlsx'),
                            "rb")
                content = file.read()
                file.close()
                return HttpResponse(content, status=200, headers={'Content-Type': 'application/vnd.ms-excel',
                                                                  'Content-Disposition': 'attachment; filename="excel_template.xlsx"'})
            else:
                return None

    class Migrate(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        knowledge_id = serializers.UUIDField(required=True, label=_('knowledge id'))
        target_knowledge_id = serializers.UUIDField(required=True, label=_('target knowledge id'))
        document_id_list = serializers.ListField(required=True, label=_('document list'),
                                                 child=serializers.UUIDField(required=True, label=_('document id')))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get('workspace_id')
            query_set = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))
            query_set = QuerySet(Knowledge).filter(id=self.data.get('target_knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))

        @transaction.atomic
        def migrate(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            knowledge_id = self.data.get('knowledge_id')
            target_knowledge_id = self.data.get('target_knowledge_id')
            knowledge = QuerySet(Knowledge).filter(id=knowledge_id).first()
            target_knowledge = QuerySet(Knowledge).filter(id=target_knowledge_id).first()
            document_id_list = self.data.get('document_id_list')
            document_list = QuerySet(Document).filter(knowledge_id=knowledge_id, id__in=document_id_list)
            paragraph_list = QuerySet(Paragraph).filter(knowledge_id=knowledge_id, document_id__in=document_id_list)

            problem_paragraph_mapping_list = QuerySet(ProblemParagraphMapping).filter(paragraph__in=paragraph_list)
            problem_list = QuerySet(Problem).filter(
                id__in=[problem_paragraph_mapping.problem_id for problem_paragraph_mapping in
                        problem_paragraph_mapping_list])
            target_problem_list = list(
                QuerySet(Problem).filter(content__in=[problem.content for problem in problem_list],
                                         knowledge_id=target_knowledge_id))
            target_handle_problem_list = [
                self.get_target_knowledge_problem(target_knowledge_id, problem_paragraph_mapping,
                                                  problem_list, target_problem_list) for
                problem_paragraph_mapping
                in
                problem_paragraph_mapping_list]

            create_problem_list = [problem for problem, is_create in target_handle_problem_list if
                                   is_create is not None and is_create]
            # 插入问题
            QuerySet(Problem).bulk_create(create_problem_list)
            # 修改mapping
            QuerySet(ProblemParagraphMapping).bulk_update(problem_paragraph_mapping_list,
                                                          ['problem_id', 'knowledge_id'])
            # 修改文档
            if knowledge.type == KnowledgeType.BASE.value and target_knowledge.type == KnowledgeType.WEB.value:
                document_list.update(knowledge_id=target_knowledge_id, type=KnowledgeType.WEB,
                                     meta={'source_url': '', 'selector': ''})
            elif target_knowledge.type == KnowledgeType.BASE.value and knowledge.type == KnowledgeType.WEB.value:
                document_list.update(knowledge_id=target_knowledge_id, type=KnowledgeType.BASE,
                                     meta={})
            else:
                document_list.update(knowledge_id=target_knowledge_id)
            model_id = None
            if knowledge.embedding_model_id != target_knowledge.embedding_model_id:
                model_id = get_embedding_model_id_by_knowledge_id(target_knowledge_id)

            pid_list = [paragraph.id for paragraph in paragraph_list]
            # 修改段落信息
            paragraph_list.update(knowledge_id=target_knowledge_id)
            # 修改向量信息
            if model_id:
                delete_embedding_by_paragraph_ids(pid_list)
                ListenerManagement.update_status(QuerySet(Document).filter(id__in=document_id_list),
                                                 TaskType.EMBEDDING,
                                                 State.PENDING)
                ListenerManagement.update_status(QuerySet(Paragraph).filter(document_id__in=document_id_list),
                                                 TaskType.EMBEDDING,
                                                 State.PENDING)
                ListenerManagement.get_aggregation_document_status_by_query_set(
                    QuerySet(Document).filter(id__in=document_id_list))()
                embedding_by_document_list.delay(document_id_list, model_id)
            else:
                update_embedding_knowledge_id(pid_list, target_knowledge_id)

        @staticmethod
        def get_target_knowledge_problem(target_knowledge_id: str,
                                         problem_paragraph_mapping,
                                         source_problem_list,
                                         target_problem_list):
            source_problem_list = [source_problem for source_problem in source_problem_list if
                                   source_problem.id == problem_paragraph_mapping.problem_id]
            problem_paragraph_mapping.knowledge_id = target_knowledge_id
            if len(source_problem_list) > 0:
                problem_content = source_problem_list[-1].content
                problem_list = [problem for problem in target_problem_list if problem.content == problem_content]
                if len(problem_list) > 0:
                    problem = problem_list[-1]
                    problem_paragraph_mapping.problem_id = problem.id
                    return problem, False
                else:
                    problem = Problem(id=uuid.uuid7(), knowledge_id=target_knowledge_id, content=problem_content)
                    target_problem_list.append(problem)
                    problem_paragraph_mapping.problem_id = problem.id
                    return problem, True
            return None

    class Query(serializers.Serializer):
        # 知识库id
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        knowledge_id = serializers.UUIDField(required=True, label=_('knowledge id'))
        name = serializers.CharField(
            required=False, max_length=128, min_length=1, allow_null=True, allow_blank=True, label=_('document name')
        )
        hit_handling_method = serializers.CharField(
            required=False, label=_('hit handling method'), allow_null=True, allow_blank=True
        )
        is_active = serializers.BooleanField(required=False, label=_('document is active'), allow_null=True)
        task_type = serializers.IntegerField(required=False, label=_('task type'))
        status = serializers.CharField(required=False, label=_('status'), allow_null=True, allow_blank=True)
        order_by = serializers.CharField(required=False, label=_('order by'), allow_null=True, allow_blank=True)

        def get_query_set(self):
            query_set = QuerySet(model=Document)
            query_set = query_set.filter(**{'knowledge_id': self.data.get("knowledge_id")})
            if 'name' in self.data and self.data.get('name') is not None:
                query_set = query_set.filter(**{'name__icontains': self.data.get('name')})
            if 'hit_handling_method' in self.data and self.data.get('hit_handling_method') not in [None, '']:
                query_set = query_set.filter(**{'hit_handling_method': self.data.get('hit_handling_method')})
            if 'is_active' in self.data and self.data.get('is_active') is not None:
                query_set = query_set.filter(**{'is_active': self.data.get('is_active')})
            if 'status' in self.data and self.data.get('status') is not None:
                task_type = self.data.get('task_type')
                status = self.data.get('status')
                if task_type is not None:
                    query_set = query_set.annotate(
                        reversed_status=Reverse('status'),
                        task_type_status=Substr('reversed_status', TaskType(task_type).value, 1),
                    ).filter(
                        task_type_status=State(status).value
                    ).values('id')
                else:
                    if status != State.SUCCESS.value:
                        query_set = query_set.filter(status__icontains=status)
                    else:
                        query_set = query_set.filter(status__iregex='^[2n]*$')
            order_by = self.data.get('order_by', '')
            order_by_query_set = QuerySet(model=get_dynamics_model(
                {'char_length': models.CharField(), 'paragraph_count': models.IntegerField(),
                 "update_time": models.IntegerField(), 'create_time': models.DateTimeField()}))
            if order_by:
                order_by_query_set = order_by_query_set.order_by(order_by)
            else:
                order_by_query_set = order_by_query_set.order_by('-create_time', 'id')
            return {
                'document_custom_sql': query_set,
                'order_by_query': order_by_query_set
            }

        def list(self):
            self.is_valid(raise_exception=True)
            query_set = self.get_query_set()
            return native_search(query_set, select_string=get_file_content(
                os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_document.sql')))

        def page(self, current_page, page_size):
            self.is_valid(raise_exception=True)
            query_set = self.get_query_set()
            return native_page_search(current_page, page_size, query_set, select_string=get_file_content(
                os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_document.sql')))

    class Sync(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('workspace id'))
        knowledge_id = serializers.UUIDField(required=False, label=_('knowledge id'))
        document_id = serializers.UUIDField(required=True, label=_('document id'))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get('workspace_id')
            query_set = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))
            document_id = self.data.get('document_id')
            first = QuerySet(Document).filter(id=document_id).first()
            if first is None:
                raise AppApiException(500, _('document id not exist'))
            if first.type != KnowledgeType.WEB:
                raise AppApiException(500, _('Synchronization is only supported for web site types'))

        @transaction.atomic
        def sync(self, with_valid=True, with_embedding=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document_id = self.data.get('document_id')
            document = QuerySet(Document).filter(id=document_id).first()
            state = State.SUCCESS
            if document.type != KnowledgeType.WEB:
                return True
            try:
                ListenerManagement.update_status(QuerySet(Document).filter(id=document_id),
                                                 TaskType.SYNC,
                                                 State.PENDING)
                ListenerManagement.get_aggregation_document_status(document_id)()
                source_url = document.meta.get('source_url')
                selector_list = document.meta.get('selector').split(
                    " ") if 'selector' in document.meta and document.meta.get('selector') is not None else []
                result = Fork(source_url, selector_list).fork()
                if result.status == 200:
                    # 删除段落
                    QuerySet(model=Paragraph).filter(document_id=document_id).delete()
                    # 删除问题
                    QuerySet(model=ProblemParagraphMapping).filter(document_id=document_id).delete()
                    delete_problems_and_mappings([document_id])
                    # 删除向量库
                    delete_embedding_by_document(document_id)
                    paragraphs = get_split_model('web.md').parse(result.content)
                    char_length = reduce(lambda x, y: x + y,
                                         [len(p.get('content')) for p in paragraphs],
                                         0)
                    QuerySet(Document).filter(id=document_id).update(char_length=char_length)
                    document_paragraph_model = DocumentSerializers.Create.get_paragraph_model(document, paragraphs)

                    paragraph_model_list = document_paragraph_model.get('paragraph_model_list')
                    problem_paragraph_object_list = document_paragraph_model.get('problem_paragraph_object_list')
                    problem_model_list, problem_paragraph_mapping_list = ProblemParagraphManage(
                        problem_paragraph_object_list, document.knowledge_id).to_problem_model_list()
                    # 批量插入段落
                    if len(paragraph_model_list) > 0:
                        max_position = Paragraph.objects.filter(document_id=document_id).aggregate(
                            max_position=Max('position')
                        )['max_position'] or 0
                        for i, paragraph in enumerate(paragraph_model_list):
                            paragraph.position = max_position + i + 1
                        QuerySet(Paragraph).bulk_create(paragraph_model_list)
                    # 批量插入问题
                    QuerySet(Problem).bulk_create(problem_model_list) if len(problem_model_list) > 0 else None
                    # 插入关联问题
                    QuerySet(ProblemParagraphMapping).bulk_create(problem_paragraph_mapping_list) if len(
                        problem_paragraph_mapping_list) > 0 else None
                    # 向量化
                    if with_embedding:
                        embedding_model_id = get_embedding_model_id_by_knowledge_id(document.knowledge_id)
                        ListenerManagement.update_status(QuerySet(Document).filter(id=document_id),
                                                         TaskType.EMBEDDING,
                                                         State.PENDING)
                        ListenerManagement.update_status(QuerySet(Paragraph).filter(document_id=document_id),
                                                         TaskType.EMBEDDING,
                                                         State.PENDING)
                        ListenerManagement.get_aggregation_document_status(document_id)()
                        embedding_by_document.delay(document_id, embedding_model_id)

                else:
                    state = State.FAILURE
            except Exception as e:
                maxkb_logger.error(f'{str(e)}:{traceback.format_exc()}')
                state = State.FAILURE
            ListenerManagement.update_status(
                QuerySet(Document).filter(id=document_id),
                TaskType.SYNC,
                state
            )
            ListenerManagement.update_status(
                QuerySet(Paragraph).filter(document_id=document_id),
                TaskType.SYNC,
                state
            )
            ListenerManagement.get_aggregation_document_status(document_id)()
            return True

    class Operate(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('workspace id'), allow_blank=True)
        document_id = serializers.UUIDField(required=True, label=_('document id'))
        knowledge_id = serializers.UUIDField(required=True, label=_('knowledge id'))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get('workspace_id')
            query_set = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))
            document_id = self.data.get('document_id')
            if not QuerySet(Document).filter(id=document_id).exists():
                raise AppApiException(500, _('document id not exist'))

        def export(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document = QuerySet(Document).filter(id=self.data.get("document_id")).first()
            paragraph_list = native_search(QuerySet(Paragraph).filter(document_id=self.data.get("document_id")),
                                           get_file_content(
                                               os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql',
                                                            'list_paragraph_document_name.sql')))
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(document_id=self.data.get("document_id")), get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_problem_mapping.sql')),
                with_table_name=True)
            data_dict, document_dict = self.merge_problem(paragraph_list, problem_mapping_list, [document])
            workbook = self.get_workbook(data_dict, document_dict)
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="data.xlsx"'
            workbook.save(response)
            return response

        def export_zip(self, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document = QuerySet(Document).filter(id=self.data.get("document_id")).first()
            paragraph_list = native_search(QuerySet(Paragraph).filter(document_id=self.data.get("document_id")),
                                           get_file_content(
                                               os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql',
                                                            'list_paragraph_document_name.sql')))
            problem_mapping_list = native_search(
                QuerySet(ProblemParagraphMapping).filter(document_id=self.data.get("document_id")), get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_problem_mapping.sql')),
                with_table_name=True)
            data_dict, document_dict = self.merge_problem(paragraph_list, problem_mapping_list, [document])
            res = [parse_image(paragraph.get('content')) for paragraph in paragraph_list]

            workbook = DocumentSerializers.Operate.get_workbook(data_dict, document_dict)
            response = HttpResponse(content_type='application/zip')
            response['Content-Disposition'] = 'attachment; filename="archive.zip"'
            zip_buffer = io.BytesIO()
            with TemporaryDirectory() as tempdir:
                knowledge_file = os.path.join(tempdir, 'knowledge.xlsx')
                workbook.save(knowledge_file)
                for r in res:
                    write_image(tempdir, r)
                zip_dir(tempdir, zip_buffer)
            response.write(zip_buffer.getvalue())
            return response

        def download_source_file(self):
            self.is_valid(raise_exception=True)
            file = QuerySet(File).filter(source_id=self.data.get('document_id')).first()
            if not file:
                raise AppApiException(500, _('File not exist. Only manually uploaded documents are supported'))
            return FileSerializer.Operate(data={'id': file.id}).get(with_valid=True)

        def one(self, with_valid=False):
            self.is_valid(raise_exception=True)
            query_set = QuerySet(model=Document)
            query_set = query_set.filter(**{'id': self.data.get("document_id")})
            return native_search({
                'document_custom_sql': query_set,
                'order_by_query': QuerySet(Document).order_by('-create_time', 'id')
            }, select_string=get_file_content(
                os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_document.sql')), with_search_one=True)

        def edit(self, instance: Dict, with_valid=False):
            if with_valid:
                self.is_valid(raise_exception=True)
            _document = QuerySet(Document).get(id=self.data.get("document_id"))
            if with_valid:
                DocumentEditInstanceSerializer(data=instance).is_valid(document=_document)
            update_keys = ['name', 'is_active', 'hit_handling_method', 'directly_return_similarity', 'meta']
            for update_key in update_keys:
                if update_key in instance and instance.get(update_key) is not None:
                    _document.__setattr__(update_key, instance.get(update_key))
            _document.save()
            return self.one()

        def cancel(self, instance, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                CancelInstanceSerializer(data=instance).is_valid()
            document_id = self.data.get("document_id")
            ListenerManagement.update_status(
                QuerySet(Paragraph).annotate(
                    reversed_status=Reverse('status'),
                    task_type_status=Substr('reversed_status', TaskType(instance.get('type')).value, 1),
                ).filter(
                    task_type_status__in=[State.PENDING.value, State.STARTED.value]
                ).filter(
                    document_id=document_id
                ).values('id'),
                TaskType(instance.get('type')),
                State.REVOKE
            )
            ListenerManagement.update_status(
                QuerySet(Document).annotate(
                    reversed_status=Reverse('status'),
                    task_type_status=Substr('reversed_status', TaskType(instance.get('type')).value,
                                            1),
                ).filter(
                    task_type_status__in=[State.PENDING.value, State.STARTED.value]
                ).filter(
                    id=document_id
                ).values('id'),
                TaskType(instance.get('type')),
                State.REVOKE
            )
            return True

        @transaction.atomic
        def delete(self):
            self.is_valid(raise_exception=True)
            document_id = self.data.get("document_id")
            source_file_ids = [
                doc['meta'].get(
                    'source_file_id'
                ) for doc in Document.objects.filter(id=document_id).values("meta")
            ]
            QuerySet(File).filter(id__in=source_file_ids).delete()
            QuerySet(File).filter(source_id=document_id, source_type=FileSourceType.DOCUMENT).delete()
            # 删除段落
            QuerySet(model=Paragraph).filter(document_id=document_id).delete()
            # 删除问题
            delete_problems_and_mappings([document_id])
            # 删除向量库
            delete_embedding_by_document(document_id)
            QuerySet(model=Document).filter(id=document_id).delete()
            return True

        def refresh(self, state_list=None, with_valid=True):
            if state_list is None:
                state_list = [State.PENDING.value, State.STARTED.value, State.SUCCESS.value, State.FAILURE.value,
                              State.REVOKE.value,
                              State.REVOKED.value, State.IGNORED.value]
            if with_valid:
                self.is_valid(raise_exception=True)
            knowledge = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id')).first()
            embedding_model_id = knowledge.embedding_model_id
            knowledge_user_id = knowledge.user_id
            embedding_model = QuerySet(Model).filter(id=embedding_model_id).first()
            if embedding_model is None:
                raise AppApiException(500, _('Model does not exist'))
            document_id = self.data.get("document_id")
            ListenerManagement.update_status(
                QuerySet(Document).filter(id=document_id), TaskType.EMBEDDING, State.PENDING
            )
            ListenerManagement.update_status(
                QuerySet(Paragraph).annotate(
                    reversed_status=Reverse('status'),
                    task_type_status=Substr('reversed_status', TaskType.EMBEDDING.value, 1),
                ).filter(task_type_status__in=state_list, document_id=document_id).values('id'),
                TaskType.EMBEDDING,
                State.PENDING
            )
            ListenerManagement.get_aggregation_document_status(document_id)()

            try:
                embedding_by_document.delay(document_id, embedding_model_id, state_list)
            except AlreadyQueued as e:
                raise AppApiException(500, _('The task is being executed, please do not send it repeatedly.'))

        @staticmethod
        def get_workbook(data_dict, document_dict):
            # 创建工作簿对象
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            if len(data_dict.keys()) == 0:
                data_dict['sheet'] = []
            for sheet_id in data_dict:
                # 添加工作表
                worksheet = workbook.create_sheet(document_dict.get(sheet_id))
                data = [
                    [gettext('Section title (optional)'),
                     gettext('Section content (required, question answer, no more than 4096 characters)'),
                     gettext('Question (optional, one per line in the cell)')],
                    *data_dict.get(sheet_id, [])
                ]
                # 写入数据到工作表
                for row_idx, row in enumerate(data):
                    for col_idx, col in enumerate(row):
                        cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                        if isinstance(col, str):
                            col = re.sub(ILLEGAL_CHARACTERS_RE, '', col)
                            if col.startswith(('=', '+', '-', '@')):
                                col = '\ufeff' + col
                        cell.value = col
                    # 创建HttpResponse对象返回Excel文件
            return workbook

        @staticmethod
        def merge_problem(paragraph_list: List[Dict], problem_mapping_list: List[Dict], document_list):
            result = {}
            document_dict = {}

            for paragraph in paragraph_list:
                problem_list = [problem_mapping.get('content') for problem_mapping in problem_mapping_list if
                                problem_mapping.get('paragraph_id') == paragraph.get('id')]
                document_sheet = result.get(paragraph.get('document_id'))
                document_name = DocumentSerializers.Operate.reset_document_name(paragraph.get('document_name'))
                d = document_dict.get(document_name)
                if d is None:
                    document_dict[document_name] = {paragraph.get('document_id')}
                else:
                    d.add(paragraph.get('document_id'))

                if document_sheet is None:
                    result[paragraph.get('document_id')] = [[paragraph.get('title'), paragraph.get('content'),
                                                             '\n'.join(problem_list)]]
                else:
                    document_sheet.append([paragraph.get('title'), paragraph.get('content'), '\n'.join(problem_list)])
            for document in document_list:
                if document.id not in result:
                    document_name = DocumentSerializers.Operate.reset_document_name(document.name)
                    result[document.id] = [[]]
                    d = document_dict.get(document_name)
                    if d is None:
                        document_dict[document_name] = {document.id}
                    else:
                        d.add(document.id)
            result_document_dict = {}
            for d_name in document_dict:
                for index, d_id in enumerate(document_dict.get(d_name)):
                    result_document_dict[d_id] = d_name if index == 0 else d_name + str(index)
            return result, result_document_dict

        @staticmethod
        def reset_document_name(document_name):
            if document_name is not None:
                document_name = document_name.strip()[0:29]
            if document_name is None or not Utils.valid_sheet_name(document_name):
                return "Sheet"
            return document_name.strip()

    class Create(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('workspace id'), allow_null=True)
        knowledge_id = serializers.UUIDField(required=True, label=_('document id'))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            if not QuerySet(Knowledge).filter(id=self.data.get('knowledge_id')).exists():
                raise AppApiException(10000, _('knowledge id not exist'))
            return True

        @staticmethod
        def post_embedding(result, document_id, knowledge_id):
            DocumentSerializers.Operate(
                data={'knowledge_id': knowledge_id, 'document_id': document_id}).refresh()
            return result

        @post(post_function=post_embedding)
        @transaction.atomic
        def save(self, instance: Dict, with_valid=False, **kwargs):
            if with_valid:
                DocumentInstanceSerializer(data=instance).is_valid(raise_exception=True)
                self.is_valid(raise_exception=True)
            knowledge_id = self.data.get('knowledge_id')
            document_paragraph_model = self.get_document_paragraph_model(knowledge_id, instance)

            document_model = document_paragraph_model.get('document')
            paragraph_model_list = document_paragraph_model.get('paragraph_model_list')
            problem_paragraph_object_list = document_paragraph_model.get('problem_paragraph_object_list')
            problem_model_list, problem_paragraph_mapping_list = (
                ProblemParagraphManage(problem_paragraph_object_list, knowledge_id).to_problem_model_list())
            # 插入文档
            document_model.save()
            # 批量插入段落
            if len(paragraph_model_list) > 0:
                max_position = Paragraph.objects.filter(document_id=document_model.id).aggregate(
                    max_position=Max('position')
                )['max_position'] or 0
                for i, paragraph in enumerate(paragraph_model_list):
                    paragraph.position = max_position + i + 1
                QuerySet(Paragraph).bulk_create(paragraph_model_list)
            # 批量插入问题
            QuerySet(Problem).bulk_create(problem_model_list) if len(problem_model_list) > 0 else None
            # 批量插入关联问题
            QuerySet(ProblemParagraphMapping).bulk_create(
                problem_paragraph_mapping_list
            ) if len(problem_paragraph_mapping_list) > 0 else None
            document_id = str(document_model.id)
            return (DocumentSerializers.Operate(
                data={'knowledge_id': knowledge_id, 'document_id': document_id}
            ).one(with_valid=True), document_id, knowledge_id)

        @staticmethod
        def get_paragraph_model(document_model, paragraph_list: List):
            knowledge_id = document_model.knowledge_id
            paragraph_model_dict_list = [
                ParagraphSerializers.Create(
                    data={
                        'knowledge_id': knowledge_id, 'document_id': str(document_model.id)
                    }).get_paragraph_problem_model(knowledge_id, document_model.id, paragraph)
                for paragraph in paragraph_list]

            paragraph_model_list = []
            problem_paragraph_object_list = []
            for paragraphs in paragraph_model_dict_list:
                paragraph = paragraphs.get('paragraph')
                for problem_model in paragraphs.get('problem_paragraph_object_list'):
                    problem_paragraph_object_list.append(problem_model)
                paragraph_model_list.append(paragraph)

            return {
                'document': document_model,
                'paragraph_model_list': paragraph_model_list,
                'problem_paragraph_object_list': problem_paragraph_object_list
            }

        @staticmethod
        def get_document_paragraph_model(knowledge_id, instance: Dict):
            source_meta = {'source_file_id': instance.get('source_file_id')} if instance.get('source_file_id') else {}
            meta = {**instance.get('meta'), **source_meta} if instance.get('meta') is not None else source_meta
            meta = convert_uuid_to_str(meta)

            document_model = Document(
                **{
                    'knowledge_id': knowledge_id,
                    'id': uuid.uuid7(),
                    'name': instance.get('name'),
                    'char_length': reduce(
                        lambda x, y: x + y,
                        [len(p.get('content')) for p in instance.get('paragraphs', [])],
                        0),
                    'meta': meta,
                    'type': instance.get('type') if instance.get('type') is not None else KnowledgeType.BASE
                })

            return DocumentSerializers.Create.get_paragraph_model(
                document_model,
                instance.get('paragraphs') if 'paragraphs' in instance else []
            )

        def save_web(self, instance: Dict, with_valid=True):
            if with_valid:
                DocumentWebInstanceSerializer(data=instance).is_valid(raise_exception=True)
                self.is_valid(raise_exception=True)
            knowledge_id = self.data.get('knowledge_id')
            source_url_list = instance.get('source_url_list')
            selector = instance.get('selector')
            sync_web_document.delay(knowledge_id, source_url_list, selector)

        def save_qa(self, instance: Dict, with_valid=True):
            if with_valid:
                DocumentInstanceQASerializer(data=instance).is_valid(raise_exception=True)
                self.is_valid(raise_exception=True)
            file_list = instance.get('file_list')
            document_list = flat_map([self.parse_qa_file(file) for file in file_list])
            return DocumentSerializers.Batch(data={
                'knowledge_id': self.data.get('knowledge_id'), 'workspace_id': self.data.get('workspace_id')
            }).batch_save(document_list)

        def save_table(self, instance: Dict, with_valid=True):
            if with_valid:
                DocumentInstanceTableSerializer(data=instance).is_valid(raise_exception=True)
                self.is_valid(raise_exception=True)
            file_list = instance.get('file_list')
            document_list = flat_map([self.parse_table_file(file) for file in file_list])
            return DocumentSerializers.Batch(data={
                'knowledge_id': self.data.get('knowledge_id'), 'workspace_id': self.data.get('workspace_id')
            }).batch_save(document_list)

        def parse_qa_file(self, file):
            #  保存源文件
            source_file_id = uuid.uuid7()
            source_file = File(
                id=source_file_id,
                file_name=file.name,
                source_type=FileSourceType.KNOWLEDGE,
                source_id=self.data.get('knowledge_id'),
                meta={}
            )
            source_file.save(file.read())
            file.seek(0)

            get_buffer = FileBufferHandle().get_buffer
            for parse_qa_handle in parse_qa_handle_list:
                if parse_qa_handle.support(file, get_buffer):
                    documents = parse_qa_handle.handle(file, get_buffer, self.save_image)
                    for doc in documents:
                        doc['source_file_id'] = source_file_id
                    return documents
            raise AppApiException(500, _('Unsupported file format'))

        def parse_table_file(self, file):
            #  保存源文件
            source_file_id = uuid.uuid7()
            source_file = File(
                id=source_file_id,
                file_name=file.name,
                source_type=FileSourceType.KNOWLEDGE,
                source_id=self.data.get('knowledge_id'),
                meta={}
            )
            source_file.save(file.read())
            file.seek(0)

            get_buffer = FileBufferHandle().get_buffer
            for parse_table_handle in parse_table_handle_list:
                if parse_table_handle.support(file, get_buffer):
                    documents = parse_table_handle.handle(file, get_buffer, self.save_image)
                    for doc in documents:
                        doc['source_file_id'] = source_file_id
                    return documents
            raise AppApiException(500, _('Unsupported file format'))

        def save_image(self, image_list):
            if image_list is not None and len(image_list) > 0:
                exist_image_list = [str(i.get('id')) for i in
                                    QuerySet(File).filter(id__in=[i.id for i in image_list]).values('id')]
                save_image_list = [image for image in image_list if not exist_image_list.__contains__(str(image.id))]
                save_image_list = list({img.id: img for img in save_image_list}.values())
                # save image
                for file in save_image_list:
                    file_bytes = file.meta.pop('content')
                    file.meta['knowledge_id'] = self.data.get('knowledge_id')
                    file.source_type = FileSourceType.KNOWLEDGE
                    file.source_id = self.data.get('knowledge_id')
                    file.save(file_bytes)

    class Split(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('workspace id'), allow_null=True)
        knowledge_id = serializers.UUIDField(required=True, label=_('knowledge id'))

        def is_valid(self, *, instance=None, raise_exception=True):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get('workspace_id')
            query_set = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))
            files = instance.get('file')
            knowledge = Knowledge.objects.filter(id=self.data.get('knowledge_id')).first()
            for f in files:
                if f.size > 1024 * 1024 * knowledge.file_size_limit:
                    raise AppApiException(500, _(
                        'The maximum size of the uploaded file cannot exceed {}MB'
                    ).format(knowledge.file_size_limit))

        def parse(self, instance):
            self.is_valid(instance=instance, raise_exception=True)
            DocumentSplitRequest(data=instance).is_valid(raise_exception=True)

            file_list = instance.get("file")
            return reduce(
                lambda x, y: [*x, *y],
                [self.file_to_paragraph(
                    f,
                    instance.get("patterns", None),
                    instance.get("with_filter", None),
                    instance.get("limit", 4096)
                ) for f in file_list],
                []
            )

        def save_image(self, image_list):
            if image_list is not None and len(image_list) > 0:
                exist_image_list = [str(i.get('id')) for i in
                                    QuerySet(File).filter(id__in=[i.id for i in image_list]).values('id')]
                save_image_list = [image for image in image_list if not exist_image_list.__contains__(str(image.id))]
                save_image_list = list({img.id: img for img in save_image_list}.values())
                # save image
                for file in save_image_list:
                    file_bytes = file.meta.pop('content')
                    file.meta['knowledge_id'] = self.data.get('knowledge_id')
                    file.source_type = FileSourceType.KNOWLEDGE
                    file.source_id = self.data.get('knowledge_id')
                    file.save(file_bytes)

        def file_to_paragraph(self, file, pattern_list: List, with_filter: bool, limit: int):
            # 保存源文件
            file_id = uuid.uuid7()
            raw_file = File(
                id=file_id,
                file_name=file.name,
                file_size=file.size,
                source_type=FileSourceType.KNOWLEDGE,
                source_id=self.data.get('knowledge_id'),
            )
            raw_file.save(file.read())
            file.seek(0)

            get_buffer = FileBufferHandle().get_buffer
            for split_handle in split_handles:
                if split_handle.support(file, get_buffer):
                    result = split_handle.handle(file, pattern_list, with_filter, limit, get_buffer, self.save_image)
                    if isinstance(result, list):
                        for item in result:
                            item['source_file_id'] = file_id
                        return result
                    result['source_file_id'] = file_id
                    return [result]
            result = default_split_handle.handle(file, pattern_list, with_filter, limit, get_buffer, self.save_image)
            if isinstance(result, list):
                for item in result:
                    item['source_file_id'] = file_id
                return result
            result['source_file_id'] = file_id
            return [result]

    class SplitPattern(serializers.Serializer):
        workspace_id = serializers.CharField(required=False, label=_('workspace id'), allow_null=True)
        knowledge_id = serializers.UUIDField(required=True, label=_('knowledge id'))

        @staticmethod
        def list():
            return [
                {'key': "#", 'value': '(?<=^)# .*|(?<=\\n)# .*'},
                {'key': '##', 'value': '(?<=\\n)(?<!#)## (?!#).*|(?<=^)(?<!#)## (?!#).*'},
                {'key': '###', 'value': "(?<=\\n)(?<!#)### (?!#).*|(?<=^)(?<!#)### (?!#).*"},
                {'key': '####', 'value': "(?<=\\n)(?<!#)#### (?!#).*|(?<=^)(?<!#)#### (?!#).*"},
                {'key': '#####', 'value': "(?<=\\n)(?<!#)##### (?!#).*|(?<=^)(?<!#)##### (?!#).*"},
                {'key': '######', 'value': "(?<=\\n)(?<!#)###### (?!#).*|(?<=^)(?<!#)###### (?!#).*"},
                {'key': '-', 'value': '(?<! )- .*'},
                {'key': _('space'), 'value': '(?<! ) (?! )'},
                {'key': _('semicolon'), 'value': '(?<!；)；(?!；)'}, {'key': _('comma'), 'value': '(?<!，)，(?!，)'},
                {'key': _('period'), 'value': '(?<!。)。(?!。)'}, {'key': _('enter'), 'value': '(?<!\\n)\\n(?!\\n)'},
                {'key': _('blank line'), 'value': '(?<!\\n)\\n\\n(?!\\n)'}
            ]

    class Batch(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        knowledge_id = serializers.UUIDField(required=True, label=_('knowledge id'))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get('workspace_id')
            query_set = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))

        @staticmethod
        def link_file(source_file_id, document_id):
            if source_file_id is None:
                return
            source_file = QuerySet(File).filter(id=source_file_id).first()
            if source_file:
                # 获取原始文件内容
                file_content = source_file.get_bytes()

                # 创建新文件对象，复制原始文件的重要属性
                new_file = File(
                    id=uuid.uuid7(),
                    file_name=source_file.file_name,
                    file_size=source_file.file_size,
                    source_type=FileSourceType.DOCUMENT,
                    source_id=document_id,  # 更新为当前知识库ID
                    meta=source_file.meta.copy() if source_file.meta else {}
                )

                # 保存文件内容和元数据
                new_file.save(file_content)

        @staticmethod
        def post_embedding(document_list, knowledge_id, workspace_id):
            for document_dict in document_list:
                DocumentSerializers.Operate(data={
                    'knowledge_id': knowledge_id,
                    'document_id': document_dict.get('id'),
                    'workspace_id': workspace_id
                }).refresh()
            return document_list

        @post(post_function=post_embedding)
        @transaction.atomic
        def batch_save(self, instance_list: List[Dict], with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            DocumentInstanceSerializer(many=True, data=instance_list).is_valid(raise_exception=True)
            workspace_id = self.data.get("workspace_id")
            knowledge_id = self.data.get("knowledge_id")
            document_model_list = []
            paragraph_model_list = []
            problem_paragraph_object_list = []
            # 插入文档
            for document in instance_list:
                document_paragraph_dict_model = DocumentSerializers.Create.get_document_paragraph_model(
                    knowledge_id,
                    document
                )
                # 保存文档和文件的关系
                document_instance = document_paragraph_dict_model.get('document')
                self.link_file(document.get('source_file_id'), document_instance.id)
                document_model_list.append(document_instance)
                for paragraph in document_paragraph_dict_model.get('paragraph_model_list'):
                    paragraph_model_list.append(paragraph)
                for problem_paragraph_object in document_paragraph_dict_model.get('problem_paragraph_object_list'):
                    problem_paragraph_object_list.append(problem_paragraph_object)

            problem_model_list, problem_paragraph_mapping_list = (
                ProblemParagraphManage(problem_paragraph_object_list, knowledge_id).to_problem_model_list()
            )
            # 插入文档
            QuerySet(Document).bulk_create(document_model_list) if len(document_model_list) > 0 else None
            # 批量插入段落
            if len(paragraph_model_list) > 0:
                for document in document_model_list:
                    max_position = Paragraph.objects.filter(document_id=document.id).aggregate(
                        max_position=Max('position')
                    )['max_position'] or 0
                    sub_list = [p for p in paragraph_model_list if p.document_id == document.id]
                    for i, paragraph in enumerate(sub_list):
                        paragraph.position = max_position + i + 1
                    QuerySet(Paragraph).bulk_create(sub_list if len(sub_list) > 0 else [])
            # 批量插入问题
            bulk_create_in_batches(Problem, problem_model_list, batch_size=1000)
            # 批量插入关联问题
            bulk_create_in_batches(ProblemParagraphMapping, problem_paragraph_mapping_list, batch_size=1000)
            # 查询文档
            query_set = QuerySet(model=Document)
            if len(document_model_list) == 0:
                return [], knowledge_id, workspace_id
            query_set = query_set.filter(**{'id__in': [d.id for d in document_model_list]})
            return native_search(
                {
                    'document_custom_sql': query_set,
                    'order_by_query': QuerySet(Document).order_by('-create_time', 'id')
                },
                select_string=get_file_content(
                    os.path.join(PROJECT_DIR, "apps", "knowledge", 'sql', 'list_document.sql')
                ),
                with_search_one=False
            ), knowledge_id, workspace_id

        def batch_sync(self, instance: Dict, with_valid=True):
            if with_valid:
                BatchSerializer(data=instance).is_valid(model=Document, raise_exception=True)
                self.is_valid(raise_exception=True)
            # 异步同步
            work_thread_pool.submit(
                lambda doc_ids: [
                    DocumentSerializers.Sync(data={
                        'document_id': doc_id,
                        'knowledge_id': self.data.get('knowledge_id'),
                        'workspace_id': self.data.get('workspace_id')
                    }).sync() for doc_id in doc_ids
                ],
                instance.get('id_list')
            )
            return True

        @transaction.atomic
        def batch_delete(self, instance: Dict, with_valid=True):
            if with_valid:
                BatchSerializer(data=instance).is_valid(model=Document, raise_exception=True)
                self.is_valid(raise_exception=True)
            document_id_list = instance.get("id_list")
            source_file_ids = [doc['meta'].get('source_file_id') for doc in
                               Document.objects.filter(id__in=document_id_list).values("meta")]
            QuerySet(File).filter(id__in=source_file_ids).delete()
            QuerySet(Document).filter(id__in=document_id_list).delete()
            QuerySet(Paragraph).filter(document_id__in=document_id_list).delete()
            delete_problems_and_mappings(document_id_list)
            # 删除向量库
            delete_embedding_by_document_list(document_id_list)
            return True

        def batch_cancel(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
                BatchCancelInstanceSerializer(data=instance).is_valid(raise_exception=True)
            document_id_list = instance.get("id_list")
            ListenerManagement.update_status(
                QuerySet(Paragraph).annotate(
                    reversed_status=Reverse('status'),
                    task_type_status=Substr('reversed_status', TaskType(instance.get('type')).value, 1),
                ).filter(
                    task_type_status__in=[State.PENDING.value, State.STARTED.value]
                ).filter(
                    document_id__in=document_id_list
                ).values('id'),
                TaskType(instance.get('type')),
                State.REVOKE
            )
            ListenerManagement.update_status(
                QuerySet(Document).annotate(
                    reversed_status=Reverse('status'),
                    task_type_status=Substr('reversed_status', TaskType(instance.get('type')).value, 1),
                ).filter(
                    task_type_status__in=[State.PENDING.value, State.STARTED.value]
                ).filter(
                    id__in=document_id_list
                ).values('id'),
                TaskType(instance.get('type')),
                State.REVOKE
            )

        def batch_edit_hit_handling(self, instance: Dict, with_valid=True):
            if with_valid:
                BatchSerializer(data=instance).is_valid(model=Document, raise_exception=True)
                hit_handling_method = instance.get('hit_handling_method')
                if hit_handling_method is None:
                    raise AppApiException(500, _('Hit handling method is required'))
                if hit_handling_method != 'optimization' and hit_handling_method != 'directly_return':
                    raise AppApiException(500, _('The hit processing method must be directly_return|optimization'))
                self.is_valid(raise_exception=True)
            document_id_list = instance.get("id_list")
            hit_handling_method = instance.get('hit_handling_method')
            directly_return_similarity = instance.get('directly_return_similarity')
            update_dict = {'hit_handling_method': hit_handling_method}
            if directly_return_similarity is not None:
                update_dict['directly_return_similarity'] = directly_return_similarity
            QuerySet(Document).filter(id__in=document_id_list).update(**update_dict)

        def batch_refresh(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document_id_list = instance.get("id_list")
            state_list = instance.get("state_list")
            knowledge_id = self.data.get('knowledge_id')
            for document_id in document_id_list:
                try:
                    DocumentSerializers.Operate(
                        data={'knowledge_id': knowledge_id, 'document_id': document_id}).refresh(state_list)
                except AlreadyQueued as e:
                    pass

    class BatchGenerateRelated(serializers.Serializer):
        workspace_id = serializers.CharField(required=True, label=_('workspace id'))
        knowledge_id = serializers.UUIDField(required=True, label=_('knowledge id'))

        def is_valid(self, *, raise_exception=False):
            super().is_valid(raise_exception=True)
            workspace_id = self.data.get('workspace_id')
            query_set = QuerySet(Knowledge).filter(id=self.data.get('knowledge_id'))
            if workspace_id:
                query_set = query_set.filter(workspace_id=workspace_id)
            if not query_set.exists():
                raise AppApiException(500, _('Knowledge id does not exist'))

        def batch_generate_related(self, instance: Dict, with_valid=True):
            if with_valid:
                self.is_valid(raise_exception=True)
            document_id_list = instance.get("document_id_list")
            model_id = instance.get("model_id")
            prompt = instance.get("prompt")
            state_list = instance.get('state_list')
            ListenerManagement.update_status(
                QuerySet(Document).filter(id__in=document_id_list),
                TaskType.GENERATE_PROBLEM,
                State.PENDING
            )
            ListenerManagement.update_status(
                QuerySet(Paragraph).annotate(
                    reversed_status=Reverse('status'),
                    task_type_status=Substr('reversed_status', TaskType.GENERATE_PROBLEM.value,
                                            1),
                ).filter(
                    task_type_status__in=state_list, document_id__in=document_id_list
                )
                .values('id'),
                TaskType.GENERATE_PROBLEM,
                State.PENDING
            )
            ListenerManagement.get_aggregation_document_status_by_query_set(
                QuerySet(Document).filter(id__in=document_id_list))()
            try:
                for document_id in document_id_list:
                    generate_related_by_document_id.delay(document_id, model_id, prompt, state_list)
            except AlreadyQueued as e:
                pass


class FileBufferHandle:
    buffer = None

    def get_buffer(self, file):
        if self.buffer is None:
            self.buffer = file.read()
        return self.buffer
